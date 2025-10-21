
# meta_store.py — SQLite helper for PACqui keywords & notes
from __future__ import annotations
import os, sqlite3, threading
from typing import Iterable, List, Optional
from pathlib import Path

DEFAULT_DB = "index_cache.sqlite"

def _norm(p: str) -> str:
    try:
        return os.path.normcase(os.path.normpath(p))
    except Exception:
        return p

class MetaStore:
    """
    Pequeña capa de acceso a SQLite para:
      - doc_keywords(fullpath, keyword, source, created_at)
      - doc_notes(fullpath, note, updated_at)

    * Concurrency-friendly*: journal_mode=WAL, busy_timeout.
    * Rutas normalizadas para comparaciones case-insensitive (Windows friendly).
    """
    def __init__(self, db_path: Optional[str | os.PathLike[str]] = None):
        self.db_path = Path(db_path) if db_path else (Path(__file__).resolve().parent / DEFAULT_DB)
        self._lock = threading.RLock()
        self._ensure_schema()



    def delete_concept_by_slug(self, slug: str):
        row = self.get_concept_by_slug(slug)
        if row:
            self.delete_concept(row["id"])

    def get_concept(self, concept_id: int) -> dict | None:
        with self._connect() as con:
            r = con.execute("SELECT id, slug, title, body, tags FROM concepts WHERE id=?",
                            (int(concept_id),)).fetchone()
        return None if not r else {"id": r[0], "slug": r[1], "title": r[2], "body": r[3], "tags": r[4]}

    def get_concept_by_slug(self, slug: str) -> dict | None:
        with self._connect() as con:
            r = con.execute("SELECT id, slug, title, body, tags FROM concepts WHERE lower(slug)=lower(?)",
                            (slug,)).fetchone()
        return None if not r else {"id": r[0], "slug": r[1], "title": r[2], "body": r[3], "tags": r[4]}

    def _slugify(self, s: str) -> str:
        import re
        slug = re.sub(r"[^A-Za-zÁÉÍÓÚÜÑáéíóúüñ0-9]+", "-", (s or "").strip().lower()).strip("-")
        return re.sub(r"-{2,}", "-", slug) or "concepto"

    def bootstrap_concepts_from_keywords(self, limit: int | None = None) -> int:
        """Crea conceptos (si no existen) a partir de doc_keywords agrupadas por término."""
        with self._connect() as con:
            rows = con.execute("""
                SELECT lower(keyword) AS kw, COUNT(*) AS n
                FROM doc_keywords
                GROUP BY lower(keyword)
                ORDER BY n DESC, kw
            """).fetchall()
        n_created = 0
        for i, (kw, _n) in enumerate(rows, 1):
            if limit and i > limit:
                break
            slug = self._slugify(kw)
            try:
                self.upsert_concept(slug, kw.capitalize(), body="", tags=kw, aliases=[kw])
                n_created += 1
            except Exception:
                pass
        return n_created

    # ---------- low level ----------
    def _connect(self) -> sqlite3.Connection:
        conn = sqlite3.connect(str(self.db_path), check_same_thread=False)
        try:
            conn.execute("PRAGMA journal_mode=WAL")
        except Exception:
            pass
        try:
            conn.execute("PRAGMA synchronous=NORMAL")
        except Exception:
            pass
        try:
            conn.execute("PRAGMA busy_timeout=3000")
        except Exception:
            pass
        return conn

    # --- NUEVO: importar índice Excel/CSV ---

    def import_index_sheet(self, sheet_path: str, replace_mode: str = "merge",
                           progress=None, progress_every: int = 200) -> dict:
        """
        Importa un índice (XLSX o CSV) con columnas:
          - RUTA            (ruta absoluta)
          - PALABRAS CLAVE  (separadas por ; , o saltos de línea)
          - OBSERVACIONES   (texto libre)
        replace_mode: "merge" | "replace"
        progress(ev, **kw): callback opcional con ev in {"total","tick","text"}.
        """
        from pathlib import Path
        import csv

        def emit(ev, **kw):
            try:
                if progress: progress(ev, **kw)
            except Exception:
                pass

        sheet = Path(sheet_path)
        if not sheet.exists():
            raise FileNotFoundError(sheet)

        def _norm_header(s: str) -> str:
            return (s or "").strip().lower()

        def _split_kws(s: str) -> list[str]:
            if not s: return []
            s = s.replace("\n", ";").replace(",", ";")
            return [x.strip() for x in s.split(";") if x.strip()]

        stats = {"rows": 0, "docs": 0, "kws_added": 0, "notes_set": 0, "replaced_docs": 0}
        seen_docs = set()

        # --- AUTOKW: genera palabras clave a partir de la ruta si no vienen en el Excel/CSV ---
        import re as _re
        def _auto_kws_from_path(ruta: str) -> list[str]:
            try:
                base = os.path.splitext(os.path.basename(ruta))[0]
                folder = os.path.basename(os.path.dirname(ruta))
                toks = set()
                for s in (base, folder):
                    for t in _re.findall(r"[A-Za-zÁÉÍÓÚÜáéíóúüÑñ0-9]{3,}", s or ""):
                        toks.add(t.lower())
                toks = [t for t in toks if len(t) >= 3]
                # Limita a 12 términos útiles para no ensuciar la DB
                return toks[:12]
            except Exception:
                return []


        def _upsert_row(ruta: str, kws_str: str, note: str):
            nonlocal stats
            if not ruta:
                return
            ruta = os.path.normcase(os.path.normpath(ruta))

            # 1) keywords del Excel/CSV
            kws = _split_kws(kws_str)

            # 2) si no hay, AUTOKW por nombre de archivo/carpeta
            from_excel = bool(kws)
            if not kws:
                kws = _auto_kws_from_path(ruta)

            # 3) replace_mode → borra previas solo 1ª vez que aparece el doc
            if replace_mode == "replace" and ruta not in seen_docs:
                self.clear_keywords(ruta)
                stats["replaced_docs"] += 1

            # 4) inserta keywords (marca origen)
            if kws:
                src = "import" + (":excel" if from_excel else ":auto")
                stats["kws_added"] += self.add_keywords(ruta, kws, source=src, replace=False)

            # 5) observaciones
            if note and str(note).strip():
                self.set_note(ruta, str(note).strip())
                stats["notes_set"] += 1

            # 6) contadores de docs únicos
            if ruta not in seen_docs:
                stats["docs"] += 1
                seen_docs.add(ruta)


        # ---------- XLSX ----------
        if sheet.suffix.lower() == ".xlsx":
            emit("text", text="Abriendo Excel…")
            try:
                import openpyxl
            except Exception as e:
                raise RuntimeError(f"Para .xlsx necesitas openpyxl: {e}")
            wb = openpyxl.load_workbook(str(sheet), read_only=True, data_only=True)
            ws = wb.active
            # total de filas (sin cabecera)
            total = max(0, int((ws.max_row or 1) - 1))
            emit("total", total=total)
            headers = [_norm_header(str(c.value or "")) for c in next(ws.iter_rows(min_row=1, max_row=1))]
            idx = {h: i for i, h in enumerate(headers)}

            def _get(row, name):
                i = idx.get(_norm_header(name))
                if i is None: return ""
                v = row[i].value
                return "" if v is None else str(v)

            done = 0
            emit("text", text="Importando filas…")
            for row in ws.iter_rows(min_row=2, values_only=False):
                ruta = _get(row, "ruta") or _get(row, "RUTA")
                palabras = _get(row, "palabras clave")
                obs = _get(row, "observaciones")
                if ruta:
                    _upsert_row(ruta, palabras, obs)
                    stats["rows"] += 1
                    done += 1
                    if done % max(1, progress_every) == 0:
                        emit("tick", done=done)
            emit("tick", done=done)
            return stats

        # ---------- CSV ----------
        emit("text", text="Contando filas (CSV)…")
        # cuenta líneas (sin cabecera)
        with open(sheet, "r", encoding="utf-8-sig", newline="") as f:
            total = sum(1 for _ in f)
        total = max(0, total - 1)
        emit("total", total=total)

        emit("text", text="Importando filas (CSV)…")
        with open(sheet, "r", encoding="utf-8-sig", newline="") as f:
            sample = f.read(4096);
            f.seek(0)
            dialect = csv.Sniffer().sniff(sample, delimiters=";,")
            r = csv.DictReader(f, dialect=dialect)
            done = 0
            for rec in r:
                rec_l = {_norm_header(k): (v or "") for k, v in rec.items()}
                ruta = rec_l.get("ruta", "") or rec_l.get("path", "")
                palabras = rec_l.get("palabras clave", "")
                obs = rec_l.get("observaciones", "")
                if ruta:
                    _upsert_row(ruta, palabras, obs)
                    stats["rows"] += 1
                    done += 1
                    if done % max(1, progress_every) == 0:
                        emit("tick", done=done)
            emit("tick", done=done)
        return stats

    def _ensure_schema(self) -> None:
        with self._lock, self._connect() as conn:
            c = conn.cursor()

            # Base (se crea si no existe; no rompe si ya existe con más columnas)
            c.execute("""
                CREATE TABLE IF NOT EXISTS doc_keywords (
                    fullpath   TEXT NOT NULL,
                    keyword    TEXT NOT NULL,
                    source     TEXT DEFAULT '',
                    created_at TEXT DEFAULT (strftime('%Y-%m-%d %H:%M:%f','now','localtime'))
                )
            """)

            # --- MIGRACIÓN: añadir columnas que falten en BDs antiguas ---
            cols = {row[1].lower() for row in c.execute("PRAGMA table_info(doc_keywords)")}
            if "source" not in cols:
                c.execute("ALTER TABLE doc_keywords ADD COLUMN source TEXT DEFAULT ''")
            if "created_at" not in cols:
                c.execute("""ALTER TABLE doc_keywords
                             ADD COLUMN created_at TEXT
                             DEFAULT (strftime('%Y-%m-%d %H:%M:%f','now','localtime'))""")

            # Índices (case-insensitive). El UNIQUE puede fallar si ya hay duplicados → fallback no único.
            c.execute("CREATE INDEX IF NOT EXISTS idx_doc_keywords_fp ON doc_keywords(fullpath)")
            c.execute("CREATE INDEX IF NOT EXISTS idx_doc_keywords_kw ON doc_keywords(keyword)")
            try:
                c.execute("""
                    CREATE UNIQUE INDEX IF NOT EXISTS idx_doc_keywords_u_expr
                    ON doc_keywords( lower(fullpath), lower(keyword) )
                """)
            except sqlite3.OperationalError:
                # Si existen duplicados ya, al menos dejamos un índice normal para acelerar
                c.execute("CREATE INDEX IF NOT EXISTS idx_doc_keywords_fp_kw ON doc_keywords(fullpath, keyword)")


            # Observaciones
            c.execute("""
                CREATE TABLE IF NOT EXISTS doc_notes (
                    fullpath   TEXT PRIMARY KEY,
                    note       TEXT NOT NULL DEFAULT '',
                    updated_at TEXT DEFAULT (strftime('%Y-%m-%d %H:%M:%f','now','localtime'))
                )
            """)
            c.execute("CREATE INDEX IF NOT EXISTS idx_doc_notes_fp_expr ON doc_notes( lower(fullpath) )")
            conn.commit()

            # --- Fuentes ancladas manualmente (persistentes) ---
            c.execute("""
                CREATE TABLE IF NOT EXISTS pinned_sources (
                    path       TEXT PRIMARY KEY,
                    name       TEXT,
                    note       TEXT,
                    weight     REAL DEFAULT 1.0,
                    created_at TEXT DEFAULT (strftime('%Y-%m-%d %H:%M:%f','now','localtime'))
                )
            """)
            c.execute("CREATE INDEX IF NOT EXISTS idx_pinned_sources_path ON pinned_sources( lower(path) )")

            # >>> HISTORICOS + CONCEPTOS
            c.executescript("""
            CREATE TABLE IF NOT EXISTS qa_log (
              id INTEGER PRIMARY KEY,
              ts TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
              session_id TEXT,
              query TEXT NOT NULL,
              answer TEXT,
              model TEXT,
              tokens_in INTEGER,
              tokens_out INTEGER,
              took_ms INTEGER
            );

            CREATE TABLE IF NOT EXISTS qa_sources (
              qa_id INTEGER NOT NULL,
              path TEXT NOT NULL,
              name TEXT,
              note TEXT,
              score REAL,
              PRIMARY KEY (qa_id, path)
            );

            CREATE TABLE IF NOT EXISTS qa_feedback (
              qa_id INTEGER PRIMARY KEY,
              rating INTEGER CHECK (rating BETWEEN 0 AND 10),
              notes TEXT,
              updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS concepts (
              id INTEGER PRIMARY KEY,
              slug TEXT UNIQUE,
              title TEXT NOT NULL,
              body TEXT NOT NULL,
              tags TEXT,
              created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
              updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS concept_alias (
              id INTEGER PRIMARY KEY,
              concept_id INTEGER NOT NULL,
              alias TEXT NOT NULL UNIQUE
            );

            CREATE INDEX IF NOT EXISTS idx_concepts_text ON concepts(title, body);
            CREATE INDEX IF NOT EXISTS idx_qa_log_ts ON qa_log(ts);
            """)


            # <<< HISTORICOS + CONCEPTOS

            # --- Fuentes por concepto (para ranking específico) ---
            c.execute("""
                CREATE TABLE IF NOT EXISTS concept_sources (
                    concept_id INTEGER NOT NULL,
                    path       TEXT NOT NULL,
                    weight     REAL DEFAULT 1.2,
                    note       TEXT,
                    PRIMARY KEY (concept_id, path)
                )
            """)
            c.execute("CREATE INDEX IF NOT EXISTS idx_concept_sources_path ON concept_sources( lower(path) )")
            conn.commit()



    # ---------- keywords ----------
    def add_keywords(self, fullpath: str, keywords: Iterable[str], source: str = "manual", replace: bool = False) -> int:
        """Inserta keywords. Si replace=True, borra antes todas las previas del doc (case-insensitive)."""
        kws = [k.strip() for k in keywords if k and k.strip()]
        if not kws:
            return 0
        with self._lock, self._connect() as conn:
            kpath = _norm(fullpath)
            if replace:
                conn.execute("DELETE FROM doc_keywords WHERE lower(fullpath) = lower(?)", (kpath,))
            n = 0
            for kw in kws:
                try:
                    conn.execute(
                        "INSERT OR IGNORE INTO doc_keywords(fullpath, keyword, source) VALUES (?, ?, ?)",
                        (kpath, kw, source or "")
                    )
                    n += conn.total_changes and 1 or 0
                except sqlite3.IntegrityError:
                    pass
            conn.commit()
            return n

    def get_keywords(self, fullpath: str) -> List[str]:
        with self._lock, self._connect() as conn:
            kpath = _norm(fullpath)
            rows = conn.execute(
                "SELECT keyword FROM doc_keywords WHERE lower(fullpath)=lower(?) ORDER BY keyword COLLATE NOCASE",
                (kpath,)
            ).fetchall()
            return [r[0] for r in rows]

    def clear_keywords(self, fullpath: str) -> int:
        with self._lock, self._connect() as conn:
            kpath = _norm(fullpath)
            cur = conn.execute("DELETE FROM doc_keywords WHERE lower(fullpath)=lower(?)", (kpath,))
            conn.commit()
            return cur.rowcount or 0

    # ---------- notes ----------
    def set_note(self, fullpath: str, note: str) -> None:
        """Upsert case-insensitive por ruta."""
        with self._lock, self._connect() as conn:
            kpath = _norm(fullpath)
            cur = conn.execute("UPDATE doc_notes SET note=?, updated_at=CURRENT_TIMESTAMP WHERE lower(fullpath)=lower(?)",
                               (note or "", kpath))
            if cur.rowcount == 0:
                conn.execute("INSERT INTO doc_notes(fullpath, note) VALUES (?, ?)", (kpath, note or ""))
            conn.commit()

    def get_note(self, fullpath: str) -> str:
        with self._lock, self._connect() as conn:
            kpath = _norm(fullpath)
            row = conn.execute("SELECT note FROM doc_notes WHERE lower(fullpath)=lower(?)", (kpath,)).fetchone()
            return row[0] if row else ""

    def delete_note(self, fullpath: str) -> bool:
        with self._lock, self._connect() as conn:
            kpath = _norm(fullpath)
            cur = conn.execute("DELETE FROM doc_notes WHERE lower(fullpath)=lower(?)", (kpath,))
            conn.commit()
            return (cur.rowcount or 0) > 0

    # ---------- util ----------
    def search_by_keyword(self, keyword_substr: str, limit: int = 1000) -> list[tuple[str, str]]:
        """Devuelve [(fullpath, keyword)] que contengan keyword_substr (case-insensitive)."""
        q = f"%{keyword_substr.strip()}%"
        with self._lock, self._connect() as conn:
            rows = conn.execute(
                "SELECT fullpath, keyword FROM doc_keywords WHERE lower(keyword) LIKE lower(?) LIMIT ?",
                (q, int(limit or 1000))
            ).fetchall()
            return [(r[0], r[1]) for r in rows]

    # === HISTÓRICOS ===
    def log_qa(self, query: str, answer: str, model: str = "", sources: list | None = None,
               session_id: str = None, tokens_in: int = None, tokens_out: int = None, took_ms: int = None) -> int:
        with self._connect() as con:
            cur = con.cursor()
            cur.execute("""INSERT INTO qa_log(ts,session_id,query,answer,model,tokens_in,tokens_out,took_ms)
                           VALUES(CURRENT_TIMESTAMP,?,?,?,?,?,?,?)""",
                        (session_id, query, answer, model, tokens_in, tokens_out, took_ms))
            qa_id = cur.lastrowid
            if sources:
                rows = []
                for s in sources:
                    rows.append((qa_id, s.get("path", ""), s.get("name"), s.get("note"), float(s.get("score") or 0)))
                cur.executemany("""INSERT OR IGNORE INTO qa_sources(qa_id,path,name,note,score)
                                   VALUES(?,?,?,?,?)""", rows)
            con.commit()
        return int(qa_id)

    def set_feedback(self, qa_id: int, rating: int, notes: str = ""):
        with self._connect() as con:
            con.execute("""INSERT INTO qa_feedback(qa_id, rating, notes, updated_at)
                           VALUES(?,?,?,CURRENT_TIMESTAMP)
                           ON CONFLICT(qa_id) DO UPDATE SET
                             rating=excluded.rating,
                             notes=excluded.notes,
                             updated_at=CURRENT_TIMESTAMP""", (qa_id, int(rating), notes))
            con.commit()

    def list_qa(self, q: str = None, limit: int = 300):
        like = f"%{(q or '').lower()}%"
        sql = """SELECT L.id, L.ts, L.query, COALESCE(F.rating,-1) AS rating
                 FROM qa_log L
                 LEFT JOIN qa_feedback F ON F.qa_id=L.id
                 {flt}
                 ORDER BY L.ts DESC LIMIT ?"""
        flt = "" if not q else "WHERE lower(L.query) LIKE ?"
        with self._connect() as con:
            cur = con.execute(sql.format(flt=flt), ((limit,) if not q else (like, limit)))
            return [dict(id=r[0], ts=r[1], query=r[2], rating=r[3]) for r in cur.fetchall()]

    def get_qa(self, qa_id: int):
        with self._connect() as con:
            qa = con.execute("""SELECT id, ts, query, answer, model, tokens_in, tokens_out, took_ms
                                FROM qa_log WHERE id=?""", (qa_id,)).fetchone()
            src = con.execute("""SELECT path, name, note, score
                                 FROM qa_sources WHERE qa_id=? ORDER BY score DESC NULLS LAST, path""",
                              (qa_id,)).fetchall()
        return {
            "id": qa[0], "ts": qa[1], "query": qa[2], "answer": qa[3], "model": qa[4],
            "tokens_in": qa[5], "tokens_out": qa[6], "took_ms": qa[7],
            "sources": [{"path": s[0], "name": s[1], "note": s[2], "score": s[3]} for s in src]
        } if qa else None

    # === CONCEPTOS ===
    def upsert_concept(self, slug: str, title: str, body: str, tags: str = "", aliases: list[str] | None = None,
                       concept_id: int = None):
        slug = (slug or title).strip().lower().replace(" ", "-")
        with self._connect() as con:
            cur = con.cursor()
            if concept_id:
                cur.execute("""UPDATE concepts SET slug=?, title=?, body=?, tags=?, updated_at=CURRENT_TIMESTAMP
                               WHERE id=?""", (slug, title, body, tags, concept_id))
                cid = concept_id
            else:
                cur.execute("""INSERT INTO concepts(slug,title,body,tags) VALUES(?,?,?,?)""",
                            (slug, title, body, tags))
                cid = cur.lastrowid
            if aliases:
                for a in aliases:
                    a = (a or "").strip().lower()
                    if not a: continue
                    try:
                        cur.execute("INSERT OR IGNORE INTO concept_alias(concept_id, alias) VALUES(?,?)", (cid, a))
                    except Exception:
                        pass
            con.commit()
        return int(cid)

    def delete_concept(self, concept_id: int):
        with self._connect() as con:
            con.execute("DELETE FROM concept_alias WHERE concept_id=?", (concept_id,))
            con.execute("DELETE FROM concepts WHERE id=?", (concept_id,))
            con.commit()

    def list_concepts(self, q: str = None, limit: int = 500):
        flt = ""
        args = []
        if q and q.strip():
            flt = "WHERE lower(title) LIKE ? OR lower(body) LIKE ? OR id IN (SELECT concept_id FROM concept_alias WHERE lower(alias) LIKE ?)"
            like = f"%{q.lower()}%";
            args = [like, like, like]
        with self._connect() as con:
            rows = con.execute(f"""SELECT id, slug, title, substr(body,1,160), tags
                                   FROM concepts {flt} ORDER BY updated_at DESC LIMIT ?""", (*args, limit)).fetchall()
        return [dict(id=r[0], slug=r[1], title=r[2], body=r[3], tags=r[4]) for r in rows]

    # ---------- concept sources (rutas ponderadas por concepto) ----------
    def list_concept_sources(self, concept_id: int) -> list[dict]:
        with self._lock, self._connect() as conn:
            rows = conn.execute(
                "SELECT path, COALESCE(weight,1.2), note FROM concept_sources WHERE concept_id=?",
                (int(concept_id),)
            ).fetchall()
        return [{"path": r[0], "weight": float(r[1] or 1.2), "note": r[2]} for r in rows]

    def save_concept_sources(self, concept_id: int, items: list[dict], replace: bool = False) -> int:
        """Upsert en lote de rutas (path, weight, note) ligadas a un concepto."""
        if not items:
            return 0
        with self._lock, self._connect() as conn:
            cur = conn.cursor()
            cid = int(concept_id)
            if replace:
                cur.execute("DELETE FROM concept_sources WHERE concept_id=?", (cid,))
            n = 0
            import os
            for it in (items or []):
                p = (it.get("path") or "").strip()
                if not p:
                    continue
                p = os.path.normcase(os.path.normpath(p))
                w = float(it.get("weight") or 1.2)
                note = it.get("note")
                cur.execute("""
                    INSERT INTO concept_sources(concept_id, path, weight, note)
                    VALUES(?,?,?,?)
                    ON CONFLICT(concept_id, path) DO UPDATE SET
                        weight=excluded.weight,
                        note=excluded.note
                """, (cid, p, w, note))
                n += 1
            conn.commit()
            return n

    def delete_concept_source(self, concept_id: int, path: str) -> int:
        with self._lock, self._connect() as conn:
            import os
            p = os.path.normcase(os.path.normpath(path or ""))
            cur = conn.execute(
                "DELETE FROM concept_sources WHERE concept_id=? AND lower(path)=lower(?)",
                (int(concept_id), p)
            )
            conn.commit()
            return cur.rowcount or 0


    # ---------- pinned sources (fuentes persistentes) ----------
    def save_pinned_sources(self, items: list[dict]) -> int:
        """Guarda/actualiza una lista de {'path','name','note','weight'}."""
        if not items: return 0
        with self._lock, self._connect() as conn:
            cur = conn.cursor()
            n = 0
            for it in items:
                p = (it.get("path") or "").strip()
                if not p: continue
                cur.execute("""
                    INSERT INTO pinned_sources(path,name,note,weight)
                    VALUES(?,?,?,COALESCE(?,1.0))
                    ON CONFLICT(path) DO UPDATE SET
                      name=excluded.name,
                      note=excluded.note,
                      weight=COALESCE(excluded.weight, pinned_sources.weight)
                """, (p, it.get("name"), it.get("note"), it.get("weight")))
                n += 1
            conn.commit()
            return n

    def clear_pinned_sources(self) -> None:
        with self._lock, self._connect() as conn:
            conn.execute("DELETE FROM pinned_sources")
            conn.commit()

    def list_pinned_sources(self) -> list[dict]:
        with self._lock, self._connect() as conn:
            rows = conn.execute("SELECT path,name,note,weight FROM pinned_sources ORDER BY created_at DESC").fetchall()
        return [{"path": r[0], "name": r[1], "note": r[2], "weight": float(r[3] or 1.0)} for r in rows]

    def count_pinned_sources(self) -> int:
        with self._lock, self._connect() as conn:
            row = conn.execute("SELECT COUNT(*) FROM pinned_sources").fetchone()
        return int(row[0] if row and row[0] else 0)

    def backfill_pinned_names(self) -> int:
        """Rellena 'name' con basename(path) cuando esté vacío."""
        import os
        with self._lock, self._connect() as conn:
            cur = conn.cursor()
            rows = cur.execute(
                "SELECT path FROM pinned_sources WHERE name IS NULL OR trim(name)=''"
            ).fetchall()
            n = 0
            for (p,) in rows:
                if not p:
                    continue
                nm = os.path.basename(p)
                n += cur.execute(
                    "UPDATE pinned_sources SET name=? WHERE path=?",
                    (nm, p)
                ).rowcount
            conn.commit()
            return n

    def delete_pinned_sources(self, paths) -> int:
        items = [(p or "").strip() for p in (paths or [])]
        if not items:
            return 0
        with self._lock, self._connect() as conn:
            cur = conn.cursor()
            n = 0
            for p in items:
                if not p:
                    continue
                n += cur.execute("DELETE FROM pinned_sources WHERE lower(path)=lower(?)", (p,)).rowcount
            conn.commit()
            return n

    def concept_context_for(self, text: str, max_chars: int = 600, top_k: int = 5) -> str:
        """Devuelve un bloque corto con los conceptos que ‘matchean’ la consulta."""
        import re
        toks = set(re.findall(r"[a-z0-9áéíóúüñ]{3,}", (text or "").lower()))
        if not toks: return ""
        like_list = [f"%{t}%" for t in toks]
        placeholders = ",".join(["?"] * len(like_list))
        sql = f"""
          SELECT c.title, c.body, c.tags
          FROM concepts c
          WHERE {" OR ".join(["lower(c.title) LIKE ?", "lower(c.body) LIKE ?"])}
             OR c.id IN (SELECT concept_id FROM concept_alias WHERE lower(alias) IN ({placeholders}))
          LIMIT {top_k}
        """
        args = []
        for t in toks:
            args += [f"%{t}%", f"%{t}%"]
        args += list(toks)
        with self._connect() as con:
            rows = con.execute(sql, args).fetchall()
        items = []
        used = 0
        for title, body, tags in rows:
            frag = (body.strip() if len(body) <= 220 else body.strip()[:220].rstrip() + "…")
            chunk = f"— {title}: {frag}"
            if tags: chunk += f"  [#{tags}]"
            if used + len(chunk) > max_chars: break
            used += len(chunk)
            items.append(chunk)
        return "\n".join(items)


if __name__ == "__main__":
    # Pequeño CLI de apoyo:
    #   python meta_store.py add-kw "C:\ruta\doc.pdf" "kw1; kw2; kw3"
    #   python meta_store.py set-note "C:\ruta\doc.pdf" "texto de observación"
    #   python meta_store.py get-kw "C:\ruta\doc.pdf"
    #   python meta_store.py get-note "C:\ruta\doc.pdf"
    import sys
    if len(sys.argv) < 3:
        print("Uso: add-kw|set-note|get-kw|get-note <fullpath> [valor]")
        raise SystemExit(2)
    cmd = sys.argv[1].lower()
    fullpath = sys.argv[2]
    store = MetaStore()
    if cmd == "add-kw":
        vals = (sys.argv[3] if len(sys.argv) > 3 else "")
        kws = [x.strip() for x in vals.replace(",", ";").split(";") if x.strip()]
        n = store.add_keywords(fullpath, kws, source="manual")
        print(f"Añadidas {n} keywords.")
    elif cmd == "set-note":
        note = sys.argv[3] if len(sys.argv) > 3 else ""
        store.set_note(fullpath, note)
        print("Nota guardada.")
    elif cmd == "get-kw":
        print("; ".join(store.get_keywords(fullpath)))
    elif cmd == "get-note":
        print(store.get_note(fullpath))
    else:
        print("Comando no reconocido.")
