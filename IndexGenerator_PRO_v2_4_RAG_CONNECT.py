# __build__: IndexGenerator_PRO_v2_FIXED3 (2025-09-17)
# -*- coding: utf-8 -*-
r"""
IndexGenerator – v3.2 (Scraper de palabras clave)
Novedades/Arreglos clave en esta revisión:
- FIX: Método faltante `_load_config` causaba AttributeError en `__init__` → ahora está presente y probado.
- FIX: Helpers SQLite (_db_*) ahora son métodos de la clase.
- FIX: Búsqueda SQLite devolvía `mod` pero la UI esperaba `mod_str` → unificado a `mod_str`.
- FIX: Menú contextual en resultados y árbol (abrir fichero/carpeta).
- FIX: El árbol pisaba los resultados de búsqueda al seleccionarse programáticamente → bandera `_suspend_dir_select`.
- MEJORA: Fallback en memoria si SQLite no devuelve filas.
- MEJORA: Exportación Excel con **CARPETA BASE** y **LOCALIZACIÓN relativa** "<BASE>\\sub\\sub".
- MEJORA: Ventana de progreso determinista + tramo indeterminado al guardar.

Requisitos opcionales:
  pip install openpyxl
"""
import os
import sqlite3
import sys
import json
import time
import re
import queue
import threading
import shutil
import tempfile
from datetime import datetime
from pathlib import Path
import hashlib
import struct

import tkinter as tk
import xml.etree.ElementTree as ET
import zipfile
import textwrap
from tkinter import font as tkfont
from tkinter import ttk, messagebox, filedialog
from tkinter.scrolledtext import ScrolledText

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell import WriteOnlyCell
import collections
import math
import zipfile
import xml.etree.ElementTree as ET


APP_NAME = "IndexGenerator"
APP_VERSION = "3.2-fast"
CONFIG_PATH = Path.home() / ".organizador_red.json"


def open_in_explorer(path: Path) -> None:
    try:
        if sys.platform.startswith("win"):
            os.startfile(str(path))
        elif sys.platform == "darwin":
            os.system(f'open "{path}"')
        else:
            os.system(f'xdg-open "{path}"')
    except Exception as e:
        messagebox.showerror(APP_NAME, f"No se pudo abrir:\n{e}")


def format_size(nbytes: int) -> str:
    step = 1024.0
    for unit in ["B", "KB", "MB", "GB", "TB"]:
        if nbytes < step:
            return f"{nbytes:.0f} {unit}" if unit == "B" else f"{nbytes:.1f} {unit}"
        nbytes /= step
    return f"{nbytes:.1f} PB"


def hex_adjust(hexcolor: str, factor: float) -> str:
    hexcolor = hexcolor.lstrip("#")
    r = int(hexcolor[0:2], 16)
    g = int(hexcolor[2:4], 16)
    b = int(hexcolor[4:6], 16)
    r = max(0, min(255, int(r * factor)))
    g = max(0, min(255, int(g * factor)))
    b = max(0, min(255, int(b * factor)))
    return f"#{r:02x}{g:02x}{b:02x}"



# ============================ LINGÜÍSTICA / NLP (mínima) ============================
SPANISH_STOPWORDS = {
    'a','acá','ahí','al','algo','algún','alguna','algunas','alguno','algunos','allá','allí','ambas','ambos',
    'ante','anteayer','antes','aquel','aquella','aquellas','aquello','aquellos','aquí','arriba','así','atrás',
    'aun','aunque','bajo','bastante','bien','cada','casi','como','con','conmigo','contigo','contra','cual',
    'cuales','cualquier','cualquiera','cuyas','cuyos','cuya','cuyo','de','dejar','del','demasiado','demás',
    'dentro','deprisa','desde','despacio','donde','dos','el','él','ella','ellas','ello','ellos','emplear',
    'en','encima','entonces','entre','era','eran','es','esa','esas','ese','eso','esos','esta','está','estaba',
    'estaban','estado','estar','estará','estas','este','esto','estos','estoy','fin','fue','fueron','fui','fuimos',
    'gracias','gran','grande','ha','haber','había','habían','hace','hacen','hacer','hacerlo','hacia','han','hasta',
    'hay','hoy','la','las','le','lo','los','luego','mal','más','me','menos','mi','mis','mío','mía','míos','mías',
    'mientras','muy','nada','nadie','ni','ningún','ninguna','ninguno','no','nos','nosotras','nosotros','nuestra',
    'nuestras','nuestro','nuestros','nunca','o','os','otra','otras','otro','otros','para','pero','poco','por',
    'porque','primero','puede','pueden','pues','qué','que','quien','quién','quienes','saber','se','según','ser',
    'si','sí','siempre','siendo','sin','sobre','sois','solamente','solo','su','sus','tal','también','tampoco',
    'tan','tanto','te','tenéis','tengo','tener','tiene','tienen','todo','todos','tras','tu','tus','tuya','tuyo',
    'tuyos','tuyas','un','una','unas','uno','unos','usted','ustedes','va','vamos','van','varias','varios','vosotras',
    'vosotros','voy','ya','yo'
}

def _simple_tokenize_es(text: str):
    if not text:
        return []
    tokens = re.findall(r"[A-Za-zÁÉÍÓÚÜáéíóúüÑñ0-9]{3,}", text, flags=re.UNICODE)
    return [t.lower() for t in tokens]

def _is_probable_noun_es(token: str):
    # Solo sustantivos (heurística). Excluimos verbos/tiempos comunes.
    if not token or len(token) < 3:
        return False
    # No números puros
    if token.isdigit():
        return False
    # Stopwords fuera
    if token in SPANISH_STOPWORDS:
        return False
    # Adverbios y derivados largos
    if token.endswith(("mente","amiento","imientos","imiento","aciones","ación","sión","mente")):
        return False
    # Gerundios y participios
    if token.endswith(("ando","iendo","yendo","ado","ada","ados","adas","ido","ida","idos","idas")):
        return False
    # Infinitivos
    if token.endswith(("ar","er","ir")):
        return False
    # Conjugaciones frecuentes (muy simple; puede sobrefiltrar algunos sustantivos cortos)
    if token.endswith(("aré","arás","ará","aremos","arán",
                       "eré","erás","erá","eremos","erán",
                       "iré","irás","irá","iremos","irán",
                       "aba","abas","aban","íamos","ías","ían",
                       "aste","aron","iste","ieron","amos","imos",
                       "aré","ería","erías","erían","iría","irías","irían")):
        return False
    # Aceptamos siglas/palabras alfanuméricas si no pasan filtros anteriores
    return True

class OrganizadorFrame(ttk.Frame):
        # =================== CONFIG (load/save) ===================
    def _config_dir(self):
        import os
        base = os.environ.get("LOCALAPPDATA") or os.path.expanduser("~")
        p = os.path.join(base, "IndexGenerator")
        try:
            os.makedirs(p, exist_ok=True)
        except Exception:
            pass
        return p

    def _config_path(self):
        import os
        return os.path.join(self._config_dir(), "config.json")

    def _load_config(self):
        import json
        self.config = {
            "last_base_path": "",
            "window_geometry": "",
            "split_pos": 320,
            "theme": "light"
        }
        try:
            with open(self._config_path(), "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                self.config.update(data)
        except Exception:
            # Si no hay config previa, seguimos con defaults
            pass

    def _save_config(self):
        import json
        try:
            with open(self._config_path(), "w", encoding="utf-8") as f:
                json.dump(self.config, f, ensure_ascii=False, indent=2)
        except Exception:
            pass
    # ================= /CONFIG =================

# ===========================
    # Scraper entrypoint + stub
    # ===========================
    def _cmd_scraper_entry(self):
        """Entrypoint robusto para el botón 'Scraper (palabras clave)'."""
        handler = getattr(self, 'cmd_scraper', None)
        if callable(handler):
            try:
                return handler()
            except Exception as e:
                try:
                    from tkinter import messagebox
                    messagebox.showerror('Scraper', f'Error ejecutando cmd_scraper:\n{e}')
                except Exception:
                    pass
        return self._cmd_scraper_stub()

    def _cmd_scraper_stub(self):
        """Stub seguro cuando el método real no está disponible."""
        try:
            import tkinter as tk
            from tkinter import ttk, messagebox
            top = tk.Toplevel(self.master if hasattr(self, 'master') else self.root)
            top.title('Scraper (palabras clave)')
            frm = ttk.Frame(top, padding=16)
            frm.pack(fill='both', expand=True)
            ttk.Label(frm, text='Scraper no disponible en esta build.', font=('Segoe UI', 11, 'bold')).pack(anchor='w')
            ttk.Label(frm, text='Este es un stub seguro. El botón no fallará aunque no exista cmd_scraper.', wraplength=420, justify='left').pack(anchor='w', pady=(8,12))
            ttk.Button(frm, text='Cerrar', command=top.destroy).pack(anchor='e')
        except Exception:
            pass

    # ===========================
    # Asistente LLM (método dentro de la clase)
    # ===========================
    def _open_llm(self, event=None):
        """Abre el chat con un modelo GGUF local (llama-cpp-python) o muestra aviso."""
        try:
            LLMChatDialog(self.master if hasattr(self, 'master') else self.root, self)
        except NameError:
            # Si LLMChatDialog no está definido, mostrar stub
            try:
                import tkinter as tk
                from tkinter import ttk, messagebox
                top = tk.Toplevel(self.master if hasattr(self, 'master') else self.root)
                top.title("PACqui — Asistente LLM (local)")
                frm = ttk.Frame(top, padding=16)
                frm.pack(fill='both', expand=True)
                ttk.Label(frm, text="PACqui (Asistente)", font=('Segoe UI', 12, 'bold')).pack(anchor='w')
                ttk.Label(frm, text="Integración LLM no disponible en esta build. Este es un stub seguro.", wraplength=420, justify='left').pack(anchor='w', pady=(8,12))
                ttk.Button(frm, text="Cerrar", command=top.destroy).pack(anchor='e')
            except Exception:
                pass
        except Exception as e:
            try:
                from tkinter import messagebox
                messagebox.showerror("Asistente LLM", f"No se pudo abrir el asistente LLM:\n{e}")
            except Exception:
                pass

    # ============================ INIT ============================
    def __init__(self, master):
        super().__init__(master)
        self.master = master

        # Estado
        self.queue = queue.Queue()
        self.cancel_event = threading.Event()
        self.base_path: Path | None = None
        self.simular_var = tk.BooleanVar(value=True)
        self.file_index: list[dict] = []        # Índice en memoria tras escaneo
        self.dir_nodes: dict[str, str] = {}     # iid -> ruta absoluta
        self._suspend_dir_select = False        # ← evita que el árbol pise resultados

        # Carga configuración y UI
        self._load_config()
        self._make_layout()
        self._update_buttons_state()

        # Eventos
        self.master.protocol("WM_DELETE_WINDOW", self._on_close)
        self.master.bind("<F5>", lambda e: self.cmd_escanear())
        self.master.bind("<Control-l>", lambda e: self._limpiar_textos())
        self.after(80, self._poll_queue)

        # Blink inicial
        self._start_blink([self.btn_sel_base_bot], base_color=self.colors["select"], duration_ms=8_000)

        # Árbol inicial
        self._build_dir_tree()
        self.after(150, self._place_sash_initial)

    

        # --- Diagnóstico RAG al arrancar ---
        try:
            has_rag = hasattr(self.__class__, "_index_file_chunks") or callable(getattr(self, "_index_file_chunks", None))
            inst_attr = getattr(self, "_index_file_chunks", None)
            cls_attr = getattr(self.__class__, "_index_file_chunks", None)
            self._append_msg(f"RAG check → inst={'OK' if callable(inst_attr) else type(inst_attr).__name__}; cls={'OK' if callable(cls_attr) else type(cls_attr).__name__}; src={__file__}", "DEBUG")
            if has_rag:
                self._append_msg("RAG activo: se indexarán fragmentos de texto en SQLite.", "INFO")
            else:
                self._append_msg("RAG desactivado: el método _index_file_chunks no está disponible en esta build.", "WARN")

        except Exception:
            pass
# ============================ LAYOUT ============================
    def _make_layout(self):
        self.columnconfigure(0, weight=1)
        self.rowconfigure(1, weight=1)

        self.colors = {
            "select": "#bde0fe", "open": "#caffbf", "scan": "#ffd6a5", "dry": "#fdffb6",
            "clear": "#ffadad", "search": "#bdb2ff", "reset": "#e9edc9",
            "excel": "#cfe8ff", "remove": "#ffcad4",
        }

        # Centro: Panel con árbol + notebook
        center = tk.PanedWindow(self, orient=tk.HORIZONTAL, sashwidth=6)
        center.grid(row=1, column=0, sticky="nsew")
        self.center = center

        # ---- Izquierda: árbol de carpetas ----
        left = ttk.Frame(center, padding=(6, 4))
        center.add(left, stretch='always')
        left.rowconfigure(1, weight=1)
        left.columnconfigure(0, weight=1)
        ttk.Label(left, text="📁 Estructura de la carpeta base", anchor="w").grid(row=0, column=0, sticky="w", pady=(0, 4))
        self.dir_tree = ttk.Treeview(left, show="tree")
        self.dir_tree.grid(row=1, column=0, sticky="nsew")
        ysb_l = ttk.Scrollbar(left, orient="vertical", command=self.dir_tree.yview)
        self.dir_tree.configure(yscroll=ysb_l.set)
        ysb_l.grid(row=1, column=1, sticky="ns")
        self.dir_tree.bind("<<TreeviewOpen>>", self._on_dir_open)
        self.dir_tree.bind("<<TreeviewSelect>>", self._on_dir_select)
        self.dir_tree.bind("<Button-3>", self._on_dir_context)

        # ---- Derecha: notebook con Resultados / Mensajes ----
        right = ttk.Frame(center, padding=(6, 4))
        center.add(right, stretch='always')
        right.rowconfigure(0, weight=1)
        right.columnconfigure(0, weight=1)
        nb = ttk.Notebook(right)
        nb.grid(row=0, column=0, sticky="nsew")

        # Pestaña: Resultados
        tab_res = ttk.Frame(nb)
        tab_res.columnconfigure(0, weight=1)
        tab_res.rowconfigure(0, weight=1)
        nb.add(tab_res, text="Resultados")
        cols = ("nombre", "ext", "tam", "modificado", "carpeta", "ruta")
        self.tree = ttk.Treeview(tab_res, columns=cols, show="headings")
        self.tree.grid(row=0, column=0, sticky="nsew")
        for c, h, w, anc in [
            ("nombre", "Nombre", 300, "w"),
            ("ext", "Ext", 70, "w"),
            ("tam", "Tamaño", 110, "e"),
            ("modificado", "Modificado", 150, "w"),
            ("carpeta", "Carpeta", 360, "w"),
            ("ruta", "Ruta completa", 520, "w"),
        ]:
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor=anc)
                # --- PanedWindow: lista (izq) / previsualización (dcha) ---
        pan = ttk.PanedWindow(tab_res, orient="horizontal")
        pan.grid(row=0, column=0, columnspan=2, sticky="nsew")
        tab_res.columnconfigure(0, weight=1)
        tab_res.rowconfigure(0, weight=1)

        # Panel izquierdo (resultados)
        pane_left = ttk.Frame(pan)
        pan.add(pane_left, weight=3)

        self.tree = ttk.Treeview(pane_left, columns=cols, show="headings")
        self.tree.grid(row=0, column=0, sticky="nsew")
        for c, h, w, anc in [
            ("nombre", "Nombre", 300, "w"),
            ("ext", "Ext", 70, "w"),
            ("tam", "Tamaño", 110, "e"),
            ("modificado", "Modificado", 150, "w"),
            ("carpeta", "Carpeta", 360, "w"),
            ("ruta", "Ruta completa", 520, "w"),
        ]:
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor=anc)
        ysb = ttk.Scrollbar(pane_left, orient="vertical", command=self.tree.yview)
        xsb = ttk.Scrollbar(pane_left, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscroll=ysb.set, xscroll=xsb.set)
        ysb.grid(row=0, column=1, sticky="ns")
        xsb.grid(row=1, column=0, sticky="ew")
        pane_left.columnconfigure(0, weight=1)
        pane_left.rowconfigure(0, weight=1)

        # Panel derecho (previsualización)
        pane_right = ttk.Frame(pan, padding=(8, 0))
        pan.add(pane_right, weight=2)

        ttk.Label(pane_right, text="Previsualización", anchor="w").grid(row=0, column=0, sticky="ew", pady=(0, 4))
        self.prev_canvas = tk.Canvas(pane_right, bd=0, highlightthickness=0, bg="#fafafa")
        self.prev_canvas.grid(row=1, column=0, sticky="nsew")
        self.prev_canvas.bind("<Configure>", lambda e: self._redraw_preview())

        # Texto de fallback (solo si no hay miniatura)
        self.prev_text = tk.Text(pane_right, wrap="word", height=3)
        self.prev_text.grid(row=2, column=0, sticky="ew", pady=(6, 0))
        self.prev_text.configure(state="disabled")

        pane_right.columnconfigure(0, weight=1)
        pane_right.rowconfigure(1, weight=1)

        self._prev_img = None
        self._prev_pil = None

        self.tree.bind("<<TreeviewSelect>>", self._on_preview_select)
        self.tree.bind("<Double-1>", self._on_open_item)
        self.tree.bind("<Button-3>", self._on_tree_context_results)

        # Posición inicial del divisor
        def _init_sash():
            try:
                total = pan.winfo_width() or tab_res.winfo_width()
                if total:
                    pan.sashpos(0, int(total * 0.7))
            except Exception:
                pass
        self.after(200, _init_sash)

        self.lbl_resumen = ttk.Label(tab_res, text="0 resultado(s)", anchor="w")
        self.lbl_resumen.grid(row=1, column=0, sticky="ew", pady=(6, 0))

        # Pestaña: Mensajes
        tab_msg = ttk.Frame(nb)
        nb.add(tab_msg, text="Mensajes")
        tab_msg.columnconfigure(0, weight=1)
        tab_msg.rowconfigure(0, weight=1)
        self.txt_msgs = ScrolledText(tab_msg, wrap="word", height=8)
        self.txt_msgs.grid(row=0, column=0, sticky="nsew")
        self.txt_msgs.tag_configure("INFO", foreground="#1f5e99")
        self.txt_msgs.tag_configure("OK", foreground="#1a7f37")
        self.txt_msgs.tag_configure("WARN", foreground="#8a6d1d")
        self.txt_msgs.tag_configure("ERR", foreground="#a61b29")
        self.txt_msgs.tag_configure("DEBUG", foreground="#555555")

        self.center.add(left, minsize=220)
        self.center.add(right, minsize=400)

        # Barra de progreso (inferior)
        prog_frame = ttk.Frame(self, padding=(10, 6, 10, 0))
        prog_frame.grid(row=2, column=0, sticky="ew")
        prog_frame.columnconfigure(1, weight=1)
        ttk.Label(prog_frame, text="Progreso:").grid(row=0, column=0, padx=(0, 8))
        self.progress = ttk.Progressbar(prog_frame, mode="determinate", maximum=100)
        self.progress.grid(row=0, column=1, sticky="ew")
        self.lbl_prog = ttk.Label(prog_frame, text="0%")
        self.lbl_prog.grid(row=0, column=2, padx=(8, 0))

        
        # --- Zona superior: AHORA EN DOS FILAS (search arriba, botones abajo) ---
        topbar = ttk.Frame(self, padding=(0, 6))
        topbar.grid(row=0, column=0, sticky="ew")
        topbar.columnconfigure(0, weight=1)

        # ===== Row 0: BÚSQUEDA =====
        row_search = ttk.Frame(topbar)
        row_search.grid(row=0, column=0, sticky="ew", pady=(0, 4))
        for i in range(8):
            row_search.columnconfigure(i, weight=0)
        row_search.columnconfigure(1, weight=1)  # Entry principal se expande

        ttk.Label(row_search, text="Texto:").grid(row=0, column=0, sticky="w")
        self.q_text = ttk.Entry(row_search)
        self.q_text.grid(row=0, column=1, sticky="ew", padx=(6, 6))
        self.q_text.bind("<Return>", lambda e: self.cmd_buscar())

        self.btn_buscar = tk.Button(row_search, text="Buscar", command=self.cmd_buscar,
                                    bg=self.colors["search"], activebackground=self.colors["search"])
        self.btn_buscar.grid(row=0, column=2, padx=(6, 6))

        self.btn_limpiar = tk.Button(row_search, text="Limpiar filtros", command=self._limpiar_filtros,
                                     bg=self.colors["reset"], activebackground=self.colors["reset"])
        self.btn_limpiar.grid(row=0, column=3, padx=(6, 12))

        ttk.Label(row_search, text="Extensiones (csv):").grid(row=0, column=4, sticky="e")
        self.q_ext = ttk.Entry(row_search, width=18)
        self.q_ext.grid(row=0, column=5, sticky="w", padx=(6, 6))

        self.chk_en_ruta = tk.BooleanVar(value=True)
        tk.Checkbutton(row_search, text="Buscar también en ruta", variable=self.chk_en_ruta).grid(row=0, column=6, sticky="w")

        self.btn_ayuda = ttk.Button(row_search, text="Ayuda", command=self.cmd_ayuda)
        self.btn_ayuda.grid(row=0, column=7, padx=(6, 0))
        self.btn_llm = tk.Button(row_search, text="PACqui (Asistente)", command=self._open_llm, bg="#e8f5e9", activebackground="#e8f5e9")
        self.btn_llm.grid(row=0, column=8, padx=(6, 0))

        # ===== Row 1: BOTONERA =====
        row_buttons = ttk.Frame(topbar)
        row_buttons.grid(row=1, column=0, sticky="ew", pady=(0, 6))
        # a) Grupo BASE
        grp_base = tk.Frame(row_buttons)
        grp_base.grid(row=0, column=0, sticky="w")
        self.btn_sel_base_bot = tk.Button(grp_base, text="Seleccionar carpeta base...", command=self.cmd_seleccionar_base,
                                          bg=self.colors["select"], activebackground=self.colors["select"])
        self.btn_sel_base_bot.grid(row=0, column=0, padx=(0, 6))
        self.btn_abrir_bot = tk.Button(grp_base, text="Abrir carpeta base", command=self.cmd_abrir_base,
                                       bg=self.colors["open"], activebackground=self.colors["open"])
        self.btn_abrir_bot.grid(row=0, column=1, padx=(0, 6))
        self.btn_eliminar_base = tk.Button(grp_base, text="Eliminar carpeta base", command=self.cmd_eliminar_base,
                                           bg=self.colors["remove"], activebackground=self.colors["remove"])
        self.btn_eliminar_base.grid(row=0, column=2, padx=(0, 6))

        ttk.Separator(row_buttons, orient="vertical").grid(row=0, column=1, sticky="ns", padx=8)

        # b) Grupo RESULTADOS
        grp_res = tk.Frame(row_buttons)
        grp_res.grid(row=0, column=2, sticky="w")
        self.btn_escanear_bot = tk.Button(grp_res, text="Escanear (F5)", command=self.cmd_escanear,
                                          bg=self.colors["scan"], activebackground=self.colors["scan"])
        self.btn_escanear_bot.grid(row=0, column=0, padx=(0, 6))
        self.chk_dry_bot = tk.Checkbutton(grp_res, text="Simular (dry-run)", variable=self.simular_var,
                                          bg=self.colors["dry"], activebackground=self.colors["dry"], selectcolor="#ffffff")
        self.chk_dry_bot.grid(row=0, column=1, padx=(0, 6))
        self.btn_vaciar_bot = tk.Button(grp_res, text="Vaciar resultados", command=lambda: self._poblar_resultados([]),
                                        bg=self.colors["clear"], activebackground=self.colors["clear"])
        self.btn_vaciar_bot.grid(row=0, column=2, padx=(0, 6))

        ttk.Separator(row_buttons, orient="vertical").grid(row=0, column=3, sticky="ns", padx=8)

        # c) Grupo SCRAPER (extraer palabras clave)
        grp_scr = tk.Frame(row_buttons)
        grp_scr.grid(row=0, column=4, sticky="w")
        self.btn_scraper = tk.Button(
    grp_scr,
    text="Scraper (palabras clave)",
    command=self._cmd_scraper_entry,
    bg=self.colors.get("ok", "#d1f2eb"),
    activebackground=self.colors.get("ok", "#d1f2eb"),
)
        self.btn_scraper.grid(row=0, column=0, padx=(0, 6))

        ttk.Separator(row_buttons, orient="vertical").grid(row=0, column=5, sticky="ns", padx=8)

        # c) Grupo EXPORTACIÓN
        grp_exp = tk.Frame(row_buttons)
        grp_exp.grid(row=0, column=6, sticky="w")
        self.btn_export_menu = tk.Button(grp_exp, text="Exportar ▾", bg=self.colors["excel"], activebackground=self.colors["excel"])
        self.btn_export_menu.grid(row=0, column=0)
        self.btn_export_menu.bind("<Button-1>", self._show_export_menu)

        # d) Grupo AYUDA rápido (opcional) - lo mantenemos arriba como botón 'Ayuda'

        # ===== Row 2: INFO DE BASE E ÍNDICE =====
        row_info = ttk.Frame(topbar)
        row_info.grid(row=2, column=0, sticky="ew")
        row_info.columnconfigure(0, weight=1)
        row_info.columnconfigure(1, weight=1)

        self.lbl_base = ttk.Label(row_info, text=self._base_label(), anchor="w")
        self.lbl_base.grid(row=0, column=0, sticky="w", pady=(4, 0))

        self.lbl_indexinfo = ttk.Label(row_info, text="Índice: 0 archivos", anchor="e", foreground="#555")
        self.lbl_indexinfo.grid(row=0, column=1, sticky="e", pady=(4, 0))

        # Atajo global de ayuda
        self.master.bind("<F1>", self._open_help)

        # --- Fila 2: info base e índice, a la derecha ---

        row2 = 1
        self.lbl_base = ttk.Label(row_info, text=self._base_label(), anchor="w")
        self.lbl_base.grid(row=row2, column=0, columnspan=12, sticky="e", pady=(8, 0))
        self.lbl_indexinfo = ttk.Label(row_info, text="Índice: 0 archivos", anchor="e", foreground="#555")
        self.lbl_indexinfo.grid(row=row2, column=0, columnspan=12, sticky="e", pady=(8, 0), padx=(0, 220))

    # ============================ BLINK ============================
    def _start_blink(self, buttons, base_color: str, duration_ms: int = 10_000, interval_ms: int = 500):
        end_time = time.time() + duration_ms / 1000.0
        def toggle():
            if time.time() >= end_time:
                for b in buttons:
                    try:
                        b.configure(bg=base_color)
                    except Exception:
                        pass
                return
            for b in buttons:
                try:
                    b.configure(bg="#ffffff" if b.cget("bg") != "#ffffff" else base_color)
                except Exception:
                    pass
            self.after(interval_ms, toggle)
        toggle()

    # ============================ ACCIONES ============================
    def cmd_seleccionar_base(self):
        path = filedialog.askdirectory(title="Selecciona la CARPETA BASE (puede ser red/UNC)", mustexist=True)
        if path:
            self.base_path = Path(path)
            self._append_msg(f"Carpeta base establecida: {self.base_path}", "OK")
            self.lbl_base.configure(text=self._base_label())
            self._save_config()
            self._build_dir_tree()
            self._update_buttons_state()
            self._start_blink([self.btn_escanear_bot], base_color=self.colors["scan"], duration_ms=6_000)

    def cmd_abrir_base(self):
        if not self.base_path:
            messagebox.showwarning(APP_NAME, "Primero selecciona la carpeta base.")
            return
        open_in_explorer(self.base_path)

    def cmd_eliminar_base(self):
        if not self.base_path:
            return
        if not messagebox.askyesno(APP_NAME, "¿Quitar la carpeta base de la vista?\n\nEsto no borra nada del disco."):
            return
        self.base_path = None
        self._append_msg("Carpeta base eliminada de la vista.", "OK")
        self.lbl_base.configure(text=self._base_label())
        self._save_config()  # ← corregida la sangría
        self._build_dir_tree()
        self._poblar_resultados([])
        self.lbl_indexinfo.configure(text="Índice: 0 archivos")
        self._set_progress(0, "0%")
        self._update_buttons_state()

    def cmd_escanear(self):
        if not self.base_path:
            messagebox.showwarning(APP_NAME, "Selecciona primero la carpeta base (UNC en Windows: \\\\servidor\\recurso).")
            return
        self.cancel_event.clear()
        self._set_progress(0)
        self.file_index.clear()
        self._poblar_resultados([])
        self.lbl_indexinfo.configure(text="Índice: escaneando...")
        t = threading.Thread(target=self._worker_scan, daemon=True)
        t.start()
        self._append_msg("Escaneo iniciado en segundo plano...", "INFO")

    # ============================ WORKER SCAN ============================
    def _worker_scan(self):
        try:
            total_files = 0
            for _root, _dirs, files in os.walk(self.base_path):
                total_files += len(files)
            if total_files == 0:
                self.queue.put(("msg", ("No se han encontrado archivos.", "WARN")))
                self.queue.put(("progress", (0, "0%")))
                self.queue.put(("index_ready", 0))
                return

            processed = 0
            start = time.time()
            tmp_index: list[dict] = []
            self._db_open_for_scan()
            for root, _dirs, files in os.walk(self.base_path):
                for fname in files:
                    if self.cancel_event.is_set():
                        self.queue.put(("msg", ("Operación cancelada por el usuario.", "WARN")))
                        self.queue.put(("index_ready", len(tmp_index)))
                        return
                    full = Path(root) / fname
                    try:
                        stat = full.stat()
                        size = stat.st_size
                        mtime = stat.st_mtime
                        mod_str = datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M")
                    except Exception:
                        continue
                    ext = full.suffix.lower().lstrip(".") or ""
                    entry = {
                        "nombre": fname,
                        "ext": ext,
                        "tam": size,
                        "mod_ts": mtime,
                        "mod_str": mod_str,
                        "carpeta": str(full.parent),
                        "ruta": str(full),
                    }
                    tmp_index.append(entry)
                    self._db_insert_row(entry)
                    processed += 1
                    
                    # RAG index (llamada protegida y opcional)
                    has_rag = hasattr(self.__class__, "_index_file_chunks") or callable(getattr(self, "_index_file_chunks", None))
                    if has_rag:
                        try:
                            if ext in {'txt','py','md','csv','log','ini','json','xml','yaml','yml','sql','html','htm','docx','pptx','xlsx','xlsm','xltx','pdf'}:
                                self._index_file_chunks(str(full), mtime)
                        except Exception as _e:
                            self.queue.put(("msg", (f"RAG: fallo indexando {fname}: {_e}", "WARN")))
                    else:
                        # Silencia el AttributeError y deja una traza discreta
                        if ext in {'txt','py','md','csv','log','ini','json','xml','yaml','yml','sql','html','htm','docx','pptx','xlsx','xlsm','xltx','pdf'}:
                            self.queue.put(("msg", (f"RAG: saltado (método no disponible) — {fname}", "DEBUG")))

                    if processed % 50 == 0 or processed == total_files:
                        percent = int((processed / total_files) * 100)
                        self.queue.put(("progress", (percent, f"{percent}%")))
                        elapsed_now = time.time() - start
                        speed = processed / elapsed_now if elapsed_now > 0 else 0.0
                        rem = max(total_files - processed, 0)
                        eta_sec = int(rem / speed) if speed > 0 else 0
                        eta_h = eta_sec // 3600; eta_m = (eta_sec % 3600) // 60; eta_s = eta_sec % 60
                        self.queue.put(("msg", (f"Escaneo: {processed}/{total_files} ({percent}%) · ~{speed:.0f} ficheros/s · ETA {eta_h:02d}:{eta_m:02d}:{eta_s:02d}", "INFO")))

            elapsed = time.time() - start
            self._db_finalize_scan()
            self.queue.put(("msg", (f"Escaneo completado. Archivos indexados: {processed}. Tiempo: {elapsed:.2f}s", "OK")))
            self.queue.put(("index_set", tmp_index))
        except Exception as e:
            self.queue.put(("msg", (f"ERROR en escaneo: {e}", "ERR")))
            self.queue.put(("index_ready", 0))

    # ============================ SQLITE HELPERS (MÉTODOS) ============================
    def _db_path(self) -> str:
        try:
            base = os.path.dirname(os.path.abspath(__file__))
        except Exception:
            base = os.getcwd()
        return os.path.join(base, "index_cache.sqlite")
    def _db_migrate_schema(self, conn):
        """Garantiza que la tabla 'files' existe y migra columnas nuevas."""
        try:
            c = conn.cursor()
            # Crear tabla base si no existe
            c.execute("CREATE TABLE IF NOT EXISTS files(id INTEGER PRIMARY KEY, name TEXT, ext TEXT, size INTEGER, mtime_ts REAL, mtime_str TEXT, dir TEXT, fullpath TEXT)")
            # Columnas existentes
            cols = {row[1] for row in c.execute("PRAGMA table_info(files)")}
            missing = []
            if "mtime_ts" not in cols:
                c.execute("ALTER TABLE files ADD COLUMN mtime_ts REAL")
                missing.append("mtime_ts")
            if "mtime_str" not in cols:
                c.execute("ALTER TABLE files ADD COLUMN mtime_str TEXT")
                missing.append("mtime_str")
            if "size" not in cols:
                c.execute("ALTER TABLE files ADD COLUMN size INTEGER")
                missing.append("size")
            # Índices
            c.execute("CREATE INDEX IF NOT EXISTS idx_files_name ON files(name)")
            c.execute("CREATE INDEX IF NOT EXISTS idx_files_ext ON files(ext)")
            c.execute("CREATE INDEX IF NOT EXISTS idx_files_dir ON files(dir)")
            c.execute("CREATE INDEX IF NOT EXISTS idx_files_mtime ON files(mtime_ts)")
            conn.commit()
            if missing:
                self._append_msg(f"SQLite migrado: añadidas columnas {', '.join(missing)}.", "INFO")
        except Exception as ex:
            self._append_msg(f"SQLite migración fallida: {ex}", "WARN")


    def _db_open_for_scan(self):
        try:
            self._db_conn = sqlite3.connect(self._db_path(), check_same_thread=False)
            # Migrar esquema si procede
            self._db_migrate_schema(self._db_conn)
            c = self._db_conn.cursor()
            c.execute("""
                CREATE TABLE IF NOT EXISTS files(
                    id INTEGER PRIMARY KEY,
                    name TEXT,
                    ext TEXT,
                    size INTEGER,
                    mtime_ts REAL,
                    mtime_str TEXT,
                    dir TEXT,
                    fullpath TEXT
                )
            """)
            c.execute("DELETE FROM files")
            self._db_conn.commit()
        except Exception as ex:
            self._db_conn = None
            self.queue.put(("msg", (f"SQLite desactivado: {ex}", "WARN")))

    def _db_insert_row(self, e: dict):
        if getattr(self, "_db_conn", None) is None:
            return
        try:
            self._db_conn.execute(
                "INSERT INTO files(name,ext,size,mtime_ts,mtime_str,dir,fullpath) VALUES (?,?,?,?,?,?,?)",
                (e["nombre"], e["ext"], int(e["tam"] or 0), float(e["mod_ts"]), e["mod_str"], e["carpeta"], e["ruta"])
            )
        except Exception as ex:
            self.queue.put(("msg", (f"SQLite insert error: {ex}", "WARN")))

    def _db_finalize_scan(self):
        if getattr(self, "_db_conn", None) is None:
            return
        try:
            self._db_conn.commit()
            n = self._db_conn.execute("SELECT COUNT(1) FROM files").fetchone()[0]
            self.queue.put(("msg", (f"SQLite: {n} filas indexadas", "DEBUG")))
        except Exception as ex:
            self.queue.put(("msg", (f"SQLite commit error: {ex}", "WARN")))

    def _db_search(self, tokens: list[str], exts: list[str], search_in_path: bool) -> list[dict]:
        """Busca en SQLite aplicando LIKE (soporta %, ?, * y escapes con \\)."""
        if getattr(self, "_db_conn", None) is None:
            try:
                self._db_conn = sqlite3.connect(self._db_path(), check_same_thread=False)
            except Exception:
                return []

        
        def _wildcard_to_like(tok: str) -> str:
            r"""
            Convierte texto a patrón SQL LIKE.
            - '*' -> '%', '?' -> '_'
            - '%' y '_' del usuario se mantienen como comodines (NO se escapan)
            - Literales escapados con '\': \%, \_, \*, \?, \\
            """
            out = []
            i = 0
            while i < len(tok):
                ch = tok[i]
                if ch == '\\' and i + 1 < len(tok):
                    nxt = tok[i+1]
                    if nxt in ('%', '_', '\\'):
                        out.append('\\' + nxt)
                    elif nxt in ('*', '?'):
                        out.append(nxt)
                    else:
                        out.append(nxt)
                    i += 2
                    continue
                if ch == '*':
                    out.append('%')
                elif ch == '?':
                    out.append('_')
                elif ch in ('%', '_'):
                    out.append(ch)
                elif ch == '\\':
                    out.append('\\\\')
                else:
                    out.append(ch)
                i += 1
            return ''.join(out)


        c = self._db_conn.cursor()
        where = []
        params = []
        for t in tokens:
            like = _wildcard_to_like(t)
            if not any(sym in like for sym in ("%", "_")):
                like = f"%{like}%"
            if search_in_path:
                where.append("(name LIKE ? ESCAPE '\\' COLLATE NOCASE OR fullpath LIKE ? ESCAPE '\\' COLLATE NOCASE)")
                params.extend([like, like])
            else:
                where.append("name LIKE ? ESCAPE '\\' COLLATE NOCASE")
                params.append(like)
        if exts:
            placeholders = ",".join("?" for _ in exts)
            where.append(f"LOWER(ext) IN ({placeholders})")
            params.extend([e.lower().lstrip(".") for e in exts])

        sql = "SELECT name, ext, size, mtime_str, dir, fullpath FROM files"
        if where:
            sql += " WHERE " + " AND ".join(where)
        sql += " ORDER BY name"
        try:
            rows = c.execute(sql, params).fetchall()
        except Exception as ex:
            self.queue.put(("msg", (f"SQLite search error: {ex}", "WARN")))
            return []
        return [{"nombre": r[0], "ext": r[1] or "", "tam": r[2], "mod_str": r[3], "carpeta": r[4], "ruta": r[5]} for r in rows]


    # ============================ QUEUE/UI LOOP ============================
    def _poll_queue(self):
        try:
            while True:
                kind, payload = self.queue.get_nowait()
                if kind == "msg":
                    text, tag = payload
                    self._append_msg(text, tag)
                elif kind == "progress":
                    value, label = payload
                    self._set_progress(value, label)
                elif kind == "index_set":
                    self.file_index = payload
                    self.lbl_indexinfo.configure(text=f"Índice: {len(self.file_index)} archivos")
                elif kind == "index_ready":
                    n = int(payload)
                    self.lbl_indexinfo.configure(text=f"Índice: {n} archivos")
                elif kind == "task_open":
                    title, total = payload
                    self._task_open(title, int(total))
                elif kind == "task_update":
                    step, status = payload
                    self._task_update(int(step), status)
                elif kind == "task_indet":
                    (msg,) = payload
                    self._task_set_indeterminate(msg)
                elif kind == "task_close":
                    self._task_close()
        except queue.Empty:
            pass
        finally:
            self.after(80, self._poll_queue)

    # ============================ BÚSQUEDA ============================
    def cmd_ayuda(self):
        """Abre el diálogo de ayuda con pestañas."""
        try:
            HelpDialog(self.master, app_name=APP_NAME, version=APP_VERSION)
        except Exception as ex:
            self._append_msg(f"No se pudo abrir la ayuda: {ex}", "WARN")

    def cmd_buscar(self):
        if not self.file_index:
            messagebox.showinfo(APP_NAME, "No hay índice cargado. Pulsa ESCANEAR primero.")
            return
        text = self.q_text.get().strip()
        exts = [e.strip().lower().lstrip(".") for e in self.q_ext.get().split(",") if e.strip()]
        search_in_path = self.chk_en_ruta.get()
        tokens = [t for t in re.split(r"\s+", text) if t]

        wc_tokens = [t for t in tokens if ('%' in t or '?' in t)]
        plain_tokens = [t for t in tokens if ('%' not in t and '?' not in t)]
        regexes = [self._compile_wildcard(tok) for tok in wc_tokens]

        def match_mem(e: dict) -> bool:
            target = e["nombre"] + ((" " + e["ruta"]) if search_in_path else "")
            t_low = target.lower()
            for s in plain_tokens:
                if s.lower() not in t_low:
                    return False
            for rx in regexes:
                if rx and not rx.search(target):
                    return False
            if exts and e["ext"].lower() not in exts:
                return False
            return True

        results = self._db_search(tokens, exts, search_in_path)
        if not results and self.file_index:
            self._append_msg("SQLite devolvió 0 filas; haciendo fallback en memoria…", "WARN")
            results = [e for e in self.file_index if match_mem(e)]

        self._poblar_resultados(results)
        self._append_msg(f"Búsqueda: {len(results)} resultado(s).", "INFO")

    def _poblar_resultados(self, rows: list[dict]):
        for iid in self.tree.get_children():
            self.tree.delete(iid)
        for e in rows:
            self.tree.insert(
                "", "end",
                values=(e.get("nombre",""), e.get("ext",""), format_size(int(e.get("tam") or 0)),
                        e.get("mod_str",""), e.get("carpeta",""), e.get("ruta",""))
            )
        self.lbl_resumen.configure(text=f"{len(rows)} resultado(s)")

    def _limpiar_filtros(self):
        self.q_text.delete(0, tk.END)
        self.q_ext.delete(0, tk.END)
        self._poblar_resultados([])

    # ============================ ÁRBOL IZQUIERDO ============================
    def _build_dir_tree(self):
        for iid in self.dir_tree.get_children():
            self.dir_tree.delete(iid)
        self.dir_nodes.clear()

        if not self.base_path or not self.base_path.exists():
            root_iid = self.dir_tree.insert("", "end", text="— Selecciona carpeta base —", open=True)
            self.dir_nodes[root_iid] = ""
            return

        root_iid = self.dir_tree.insert("", "end", text=f"📁 {self.base_path.name}", open=True)
        self.dir_nodes[root_iid] = str(self.base_path)
        self.dir_tree.insert(root_iid, "end", text="", values=("_placeholder",))

    def _on_dir_open(self, _event):
        iid = self.dir_tree.focus()
        if not iid:
            return
        children = self.dir_tree.get_children(iid)
        if not children:
            return
        first_child = children[0]
        vals = self.dir_tree.item(first_child, "values")
        if vals and vals[0] == "_placeholder":
            self.dir_tree.delete(first_child)
            path = Path(self.dir_nodes.get(iid, ""))
            try:
                subdirs = sorted([p for p in path.iterdir() if p.is_dir()], key=lambda p: p.name.lower())
            except Exception:
                subdirs = []
            for d in subdirs:
                child_iid = self.dir_tree.insert(iid, "end", text=f"📁 {d.name}")
                self.dir_nodes[child_iid] = str(d)
                try:
                    if any(p.is_dir() for p in d.iterdir()):
                        self.dir_tree.insert(child_iid, "end", text="", values=("_placeholder",))
                except Exception:
                    pass

    def _on_dir_select(self, _event):
        # Evita que una selección programática (revelar carpeta del resultado) borre la tabla de resultados
        if getattr(self, "_suspend_dir_select", False):
            return

        sel = self.dir_tree.selection()
        if not sel:
            return
        iid = sel[0]
        path = Path(self.dir_nodes.get(iid, ""))
        if not path.exists():
            return
        rows = []
        try:
            for p in sorted(path.iterdir(), key=lambda x: (not x.is_dir(), x.name.lower())):
                if p.is_file():
                    try:
                        stat = p.stat()
                        rows.append({
                            "nombre": p.name,
                            "ext": p.suffix.lower().lstrip("."),
                            "tam": stat.st_size,
                            "mod_str": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
                            "carpeta": str(p.parent),
                            "ruta": str(p),
                        })
                    except Exception:
                        continue
        except Exception as e:
            self._append_msg(f"No se pudo listar {path}: {e}", "WARN")
            return
        self._poblar_resultados(rows)

    def _on_dir_context(self, event):
        iid = self.dir_tree.identify_row(event.y)
        if not iid:
            return
        self.dir_tree.selection_set(iid)
        ruta = self.dir_nodes.get(iid, "")
        menu = tk.Menu(self, tearoff=0)
        menu.add_command(label="Abrir carpeta", command=lambda: open_in_explorer(Path(ruta)))
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    # ============================ CONTEXTUAL RESULTADOS ============================
    def _on_tree_context_results(self, event):
        iid = self.tree.identify_row(event.y)
        if not iid:
            return
        self.tree.selection_set(iid)
        vals = self.tree.item(iid, "values")
        if not vals:
            return
        _nombre, _ext, _tam, _mod, carpeta, ruta = vals
        menu = tk.Menu(self, tearoff=0)
        menu.add_command(label="Abrir fichero", command=lambda: open_in_explorer(Path(ruta)))
        menu.add_command(label="Abrir carpeta", command=lambda: open_in_explorer(Path(carpeta)))
        menu.add_separator()
        menu.add_command(label="Ver palabras clave guardadas", command=lambda: self._show_keywords_for_file(ruta))
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    # ============================ EXCEL ============================
    def _rows_for_excel(self, export_all: bool):
        rows = []
        tv = self.tree
        if not export_all and tv.get_children() and len(self.file_index) > len(self.tree.get_children()):
            for iid in tv.get_children():
                nombre, ext, tam_str, mod_str, carpeta, ruta = tv.item(iid, "values")
                try:
                    size_val = tam_str
                except Exception:
                    size_val = tam_str
                try:
                    dt = datetime.strptime(mod_str, "%Y-%m-%d %H:%M")
                except Exception:
                    dt = None
                rows.append({"nombre": nombre, "ext": ext, "tam": size_val, "fecha": dt, "carpeta": carpeta, "ruta": ruta})
        else:
            for e in self.file_index:
                rows.append({
                    "nombre": e["nombre"],
                    "ext": e["ext"],
                    "tam": format_size(e["tam"]),
                    "fecha": datetime.fromtimestamp(e["mod_ts"]),
                    "carpeta": e["carpeta"],
                    "ruta": e["ruta"]
                })
        return rows

    @staticmethod
    def _lca_base(paths: list[Path]) -> Path:
        if not paths:
            return Path.cwd()
        parents = [str(p.parent) for p in paths]
        common = os.path.commonpath(parents)
        return Path(common)

    @staticmethod
    def _compile_wildcard(token: str):
        """Convierte %/?/* a regex (case-insensitive). Soporta escapes con \\.
        Ej.: "report\\*2024?.pdf" → "report*2024_.pdf" (SQL LIKE) y regex equivalente.
        """
        import re as _re
        if not token:
            return None
        pat = []
        i = 0
        while i < len(token):
            ch = token[i]
            if ch == '\\' and i+1 < len(token):
                # Escapado literal
                pat.append(_re.escape(token[i+1]))
                i += 2
                continue
            if ch in ("%","*"):
                pat.append(".*")
            elif ch in ("?","_"):
                pat.append(".")
            else:
                pat.append(_re.escape(ch))
            i += 1
        return _re.compile("".join(pat), _re.IGNORECASE)


    def _ensure_dir_loaded(self, parent_iid):
        """Si el nodo tiene placeholder, carga sus subcarpetas."""
        children = self.dir_tree.get_children(parent_iid)
        if not children:
            return
        first_child = children[0]
        vals = self.dir_tree.item(first_child, "values")
        if vals and vals[0] == "_placeholder":
            self.dir_tree.delete(first_child)
            path = Path(self.dir_nodes.get(parent_iid, ""))
            try:
                subdirs = sorted([p for p in path.iterdir() if p.is_dir()], key=lambda p: p.name.lower())
            except Exception:
                subdirs = []
            for d in subdirs:
                child_iid = self.dir_tree.insert(parent_iid, "end", text=f"📁 {d.name}")
                self.dir_nodes[child_iid] = str(d)
                try:
                    if any(p.is_dir() for p in d.iterdir()):
                        self.dir_tree.insert(child_iid, "end", text="", values=("_placeholder",))
                except Exception:
                    pass

    def _reveal_in_tree(self, ruta: Path):
        """Expande el árbol y selecciona la carpeta del fichero dado sin pisar resultados."""
        if not self.base_path:
            return
        base = self.base_path
        try:
            rel_parts = list(ruta.parent.relative_to(base).parts)
        except Exception:
            rel_parts = []
        base_iid = None
        for iid, p in self.dir_nodes.items():
            if p == str(base):
                base_iid = iid
                break
        if base_iid is None:
            return
        current = base_iid
        self.dir_tree.item(current, open=True)
        self._ensure_dir_loaded(current)
        path_cursor = Path(base)
        for name in rel_parts:
            path_cursor = path_cursor / name
            self._ensure_dir_loaded(current)
            match_iid = None
            for ch in self.dir_tree.get_children(current):
                if self.dir_nodes.get(ch, None) == str(path_cursor):
                    match_iid = ch
                    break
            if match_iid is None:
                for ch in self.dir_tree.get_children(current):
                    text = self.dir_tree.item(ch, "text")
                    if text.endswith(name):
                        match_iid = ch
                        break
            if match_iid is None:
                break
            current = match_iid
            self.dir_tree.item(current, open=True)
        try:
            self._suspend_dir_select = True
            self.dir_tree.selection_set(current)
            self.dir_tree.focus(current)
            self.dir_tree.see(current)
        finally:
            # reactivar después de que Tk procese la selección
            self.after(100, lambda: setattr(self, "_suspend_dir_select", False))

    def _split_dirs(self, base: Path, ruta: Path):
        """Lista de carpetas relativas a base (sin archivo)."""
        try:
            rel = ruta.relative_to(base)
        except Exception:
            return []
        return list(rel.parts[:-1])

    @staticmethod
    def _guess_version(nombre: str) -> str:
        m = re.search(r'v(?:ersi[oó]n)?[_\-\s]?(\d+(?:\.\d+)?)', nombre, flags=re.IGNORECASE)
        if not m:
            m = re.search(r'[_\-\s](\d+(?:\.\d+)?)$', nombre)
        return m.group(1) if m else ""

    @staticmethod
    def _guess_code(nombre: str) -> str:
        m = re.search(r'\b[A-Z]{2,}-[A-Za-z0-9]+', nombre)
        return m.group(0) if m else ""

    def _col_for_year(self, year: int, headers: list[str], header_to_idx: dict[str, int]) -> int:
        if year <= 2008:
            key = 'CAMBIOS EN 2008'
        elif year in (2009, 2010):
            key = 'CAMBIOS EN 2009/2010'
        elif 2011 <= year <= 2022:
            key = f'CAMBIOS EN {year}'
        elif year == 2023:
            key = 'CAMBIOS EN 2023'
        elif year == 2024:
            key = 'CAMBIOS EN 2024'
        else:
            key = 'ÚLTIMOS CAMBIOS (2025)'
        return header_to_idx.get(key, header_to_idx['ÚLTIMOS CAMBIOS (2025)'])

    def cmd_exportar_excel(self, export_all: bool):
        if not export_all and self.tree.get_children() and len(self.file_index) > len(self.tree.get_children()):
            if messagebox.askyesno(APP_NAME, f"Tienes {len(self.tree.get_children())} filas visibles pero el índice contiene {len(self.file_index)} archivos.\n\n¿Exportar TODO el índice?"):
                export_all = True

        rows = self._rows_for_excel(export_all)
        if not rows:
            messagebox.showinfo(APP_NAME, "No hay datos que exportar. Escanea o busca primero.")
            return

        default_name = f"Listado_DocuSICOP_{'TODO_' if export_all else ''}{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        save_path = filedialog.asksaveasfilename(
            title="Guardar Excel", defaultextension=".xlsx", initialfile=default_name, filetypes=[("Excel", "*.xlsx")]
        )
        if not save_path:
            return

        rutas = [Path(r["ruta"]) for r in rows]
        base = self.base_path if self.base_path else self._lca_base(rutas)
        base_name = base.name

        max_sublevels = 0
        for ruta in rutas:
            dirs = self._split_dirs(base, ruta)
            sublevels = max(0, len(dirs) - 1) if dirs else 0
            if sublevels > max_sublevels:
                max_sublevels = sublevels

        sub_headers = [f"SUBCARPETA {i}" for i in range(1, max_sublevels + 1)]
        headers = (['CARPETA BASE', 'CARPETA'] + sub_headers +
                   ['FICHERO', 'VERSIÓN', 'FECHA', 'CÓDIGO', 'CÓDIGO ANTERIOR',
                    'LOCALIZACIÓN (pendiente de actualizar)', 'FIRMAS', 'RENOVACIÓN',
                    'CAMBIOS EN 2008', 'CAMBIOS EN 2009/2010', 'CAMBIOS EN 2011', 'CAMBIOS EN 2012',
                    'CAMBIOS EN 2013', 'CAMBIOS EN 2014', 'CAMBIOS EN 2015', 'CAMBIOS EN 2016', 'CAMBIOS EN 2017',
                    'CAMBIOS EN 2018', 'CAMBIOS EN 2019', 'CAMBIOS EN 2020', 'CAMBIOS EN 2021', 'CAMBIOS EN 2022',
                    'CAMBIOS EN 2023', 'CAMBIOS EN 2024', 'ÚLTIMOS CAMBIOS (2025)', 'PENDIENTE'])

        wb = Workbook()
        ws = wb.active
        ws.title = "Listado de documentación"
        ws.append(headers)
        header_to_idx = {h: i for i, h in enumerate(headers)}
        head_fill = PatternFill("solid", fgColor="E2EFDA")
        head_font = Font(bold=True)
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=1, column=col)
            c.fill = head_fill
            c.font = head_font
            c.alignment = Alignment(wrap_text=True, vertical="center")

        self._task_open("Generando Excel", len(rows))
        for r in rows:
            ruta = Path(r["ruta"])
            dirs = self._split_dirs(base, ruta)
            carpeta = dirs[0] if dirs else ""
            subs = (dirs[1:] if len(dirs) > 1 else [])
            subs = subs + [""] * (max_sublevels - len(subs))
            fichero = ruta.name
            version = self._guess_version(ruta.stem)
            codigo = self._guess_code(ruta.stem)
            fecha = r["fecha"]

            try:
                rel_parent = ruta.parent.relative_to(base)
                rel_str = "." if str(rel_parent) == "." else str(rel_parent).replace("/", "\\")
            except Exception:
                rel_str = ""
            localizacion = base_name if rel_str in ("", ".") else f"{base_name}\\{rel_str}"

            row_vals = [""] * len(headers)
            row_vals[header_to_idx['CARPETA BASE']] = base_name
            row_vals[header_to_idx['CARPETA']] = carpeta
            for i, sh in enumerate(sub_headers):
                row_vals[header_to_idx[sh]] = subs[i] if i < len(subs) else ""
            row_vals[header_to_idx['FICHERO']] = fichero
            row_vals[header_to_idx['VERSIÓN']] = version
            row_vals[header_to_idx['FECHA']] = fecha if isinstance(fecha, datetime) else ""
            row_vals[header_to_idx['CÓDIGO']] = codigo
            row_vals[header_to_idx['CÓDIGO ANTERIOR']] = ""
            row_vals[header_to_idx['LOCALIZACIÓN (pendiente de actualizar)']] = localizacion
            row_vals[header_to_idx['FIRMAS']] = ""
            row_vals[header_to_idx['RENOVACIÓN']] = ""

            if isinstance(fecha, datetime):
                msg = f"Modificado {fecha.strftime('%Y-%m-%d %H:%M')}"
                idx = self._col_for_year(fecha.year, headers, header_to_idx)
                row_vals[idx] = msg

            ws.append(row_vals)

            c = ws.cell(row=ws.max_row, column=header_to_idx['FICHERO'] + 1)
            try:
                c.hyperlink = str(ruta)
                c.style = "Hyperlink"
            except Exception:
                pass

            for key in ['CAMBIOS EN 2008','CAMBIOS EN 2009/2010','CAMBIOS EN 2011','CAMBIOS EN 2012',
                        'CAMBIOS EN 2013','CAMBIOS EN 2014','CAMBIOS EN 2015','CAMBIOS EN 2016','CAMBIOS EN 2017',
                        'CAMBIOS EN 2018','CAMBIOS EN 2019','CAMBIOS EN 2020','CAMBIOS EN 2021','CAMBIOS EN 2022',
                        'CAMBIOS EN 2023','CAMBIOS EN 2024','ÚLTIMOS CAMBIOS (2025)','PENDIENTE']:
                ws.cell(row=ws.max_row, column=header_to_idx[key] + 1).number_format = '@'

            self._task_update(1)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

        def setw(name, width):
            col = header_to_idx.get(name)
            if col is not None:
                ws.column_dimensions[get_column_letter(col + 1)].width = width

        setw('CARPETA BASE', 26)
        setw('CARPETA', 22)
        for sh in sub_headers:
            setw(sh, 24)
        setw('FICHERO', 46)
        setw('VERSIÓN', 10)
        setw('FECHA', 16)
        setw('CÓDIGO', 20)
        setw('CÓDIGO ANTERIOR', 22)
        setw('LOCALIZACIÓN (pendiente de actualizar)', 50)

        fill_alt = PatternFill("solid", fgColor="F7F7F7")
        for row in range(2, ws.max_row + 1):
            if row % 2 == 0:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).fill = fill_alt

        self._task_set_indeterminate("Guardando archivo…")
        wb.save(save_path)
        self._task_close()
        self._append_msg(f"Excel guardado en: {save_path}", "OK")
        messagebox.showinfo(APP_NAME, f"Excel generado:\n{save_path}")

    # ============================ UTILIDAD/UI ============================
    def _append_msg(self, text: str, tag: str = "INFO"):
        self.txt_msgs.insert(tk.END, text + "\n", (tag,))
        self.txt_msgs.see(tk.END)

    def _set_progress(self, value: int, label: str | None = None):
        value = max(0, min(100, int(value)))
        self.progress["value"] = value
        self.lbl_prog.configure(text=label if label is not None else f"{value}%")
        self.update_idletasks()

    def _set_button_enabled(self, btn, enabled: bool, base_color: str):
        try:
            if enabled:
                btn.configure(state=tk.NORMAL, bg=base_color, activebackground=base_color, fg='black')
            else:
                btn.configure(state=tk.DISABLED, bg='#e5e5e5', activebackground='#e5e5e5',
                              fg='#888888', disabledforeground='#888888')
        except Exception:
            pass

    def _update_buttons_state(self):
        has_base = bool(self.base_path and Path(self.base_path).exists())
        self._set_button_enabled(self.btn_abrir_bot, has_base, self.colors['open'])
        self._set_button_enabled(self.btn_escanear_bot, has_base, self.colors['scan'])
        self._set_button_enabled(self.btn_eliminar_base, has_base, self.colors['remove'])

    def _ensure_export_menu(self):
        if getattr(self, '_export_menu', None) is not None:
            return
        m = tk.Menu(self, tearoff=0)
        m.add_command(label='Exportar Excel (visibles, rápido)', command=self.cmd_exportar_excel_rapido_visibles)
        m.add_command(label='Exportar Excel (todo, rápido 150k+)', command=self.cmd_exportar_excel_rapido)
        m.add_separator()
        m.add_command(label='Exportar Excel (visibles, clásico)', command=lambda: self.cmd_exportar_excel(False))
        m.add_command(label='Exportar Excel (todo, clásico)', command=lambda: self.cmd_exportar_excel(True))
        self._export_menu = m
    def _task_open(self, title: str, total: int):
        try:
            self._task_win.destroy()
        except Exception:
            pass
        self._task_total = max(1, int(total))
        self._task_count = 0
        win = tk.Toplevel(self)
        win.title(title)
        win.transient(self.winfo_toplevel())
        win.resizable(False, False)
        win.grab_set()
        ttk.Label(win, text=title).grid(row=0, column=0, padx=16, pady=(12, 4), sticky="w")
        pb = ttk.Progressbar(win, mode="determinate", maximum=self._task_total, length=360)
        pb.grid(row=1, column=0, padx=16, pady=(0, 8))
        lbl = ttk.Label(win, text=f"0 / {self._task_total}")
        lbl.grid(row=2, column=0, padx=16, pady=(0, 12), sticky="e")
        self._task_win, self._task_pb, self._task_lbl = win, pb, lbl
        self._task_win.update_idletasks()
    def _task_update(self, step: int = 1, status: str | None = None):
        try:
            self._task_count = min(self._task_total, self._task_count + int(step))
            self._task_pb["value"] = self._task_count
            if status:
                self._task_lbl.configure(text=status)
            else:
                self._task_lbl.configure(text=f"{self._task_count} / {self._task_total}")
            self._task_win.update_idletasks()
        except Exception:
            pass

    def _task_set_indeterminate(self, msg: str = "Procesando..."):
        try:
            self._task_lbl.configure(text=msg)
            self._task_pb.configure(mode="indeterminate")
            self._task_pb.start(60)
            self._task_win.update_idletasks()
        except Exception:
            pass
    def _task_close(self):
        try:
            self._task_pb.stop()
        except Exception:
            pass
        try:
            self._task_win.destroy()
        except Exception:
            pass
        self._task_win = None

    def _show_export_menu(self, event):
        self._ensure_export_menu()
        try:
            self._export_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self._export_menu.grab_release()

    def _limpiar_textos(self):
        self.txt_msgs.delete("1.0", tk.END)
        self._set_progress(0, "0%")
        self._poblar_resultados([])

    def _base_label(self) -> str:
        if not self.base_path:
            return "Base: [no seleccionada]"
        s = str(self.base_path)
        return "Base: " + (s if len(s) <= 80 else "..." + s[-77:])

    def _show_atajos(self):
        msg = ("Atajos:\n  • F5 → Escanear carpeta base\n  • Ctrl+L → Limpiar mensajes y resultados\n\n"
               "Consejos:\n  • Windows: rutas UNC \\\\servidor\\recurso\n  • Linux/macOS: ruta montada (/mnt/compartida)")
        messagebox.showinfo("Atajos de teclado", msg)

    def _about(self):
        messagebox.showinfo(f"Acerca de {APP_NAME}", f"{APP_NAME} v{APP_VERSION}\n\nExporta Excel jerárquico con columna de BASE y localización relativa.")

    def _on_close(self):
        self.cancel_event.set()
        self._save_config()
        self.master.destroy()

    def _on_open_item(self, _event):
        item = self.tree.focus()
        if not item:
            return
        vals = self.tree.item(item, "values")
        if not vals:
            return
        ruta = vals[5]
        try:
            open_in_explorer(Path(ruta))
        except Exception as e:
            messagebox.showerror(APP_NAME, f"No se pudo abrir:\n{e}")

    # ============================ CONFIG ============================
    def _load_config(self):
        """Carga la última carpeta base seleccionada (si existe)."""
        try:
            if CONFIG_PATH.exists():
                data = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
                base = data.get("base_path")
                if base:
                    p = Path(base)
                    if p.exists():
                        self.base_path = p
                self.llm_model_path = data.get("llm_model_path", "")
        except Exception:
            pass

    def _save_config(self):
        try:
            data = {"base_path": str(self.base_path) if self.base_path else "", "llm_model_path": getattr(self, "llm_model_path", "")}
            CONFIG_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception as e:
            self._append_msg(f"No se pudo guardar la configuración: {e}", "WARN")

    def _place_sash_initial(self):
        try:
            total = self.center.winfo_width()
            if total <= 0:
                self.after(100, self._place_sash_initial)
                return
            self.center.sash_place(0, int(total * 0.33), 1)
        except Exception:
            pass    # ---------- Previsualización
    def _on_preview_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        iid = sel[0]
        vals = self.tree.item(iid, "values")
        ruta = vals[5] if len(vals) >= 6 else None
        if ruta:
            threading.Thread(target=self._render_preview_safe, args=(ruta,), daemon=True).start()

    def _render_preview_safe(self, ruta):
        try:
            self._render_preview(ruta)
        except Exception as e:
            self.after(0, lambda: self._show_preview_text(f"Previsualización no disponible.\n\n{e}"))

    def _set_prev_text(self, text):
        text = self._normalize_text(text)
        self.prev_text.configure(state="normal")
        self.prev_text.delete("1.0", "end")
        self.prev_text.insert("1.0", text)
        self.prev_text.configure(state="disabled")

    def _show_preview_text(self, text):
        self.prev_canvas.delete("all")
        self._prev_pil = None
        self._prev_img = None
        try:
            self.prev_text.grid()
        except Exception:
            pass
        self._set_prev_text(text)

    def _show_preview_image(self, pil_img):
        # Oculta el texto y guarda la PIL para reescalado dinámico
        try:
            self.prev_text.grid_remove()
        except Exception:
            pass
        self._prev_pil = pil_img
        self._redraw_preview()

    # --- Helpers para generar miniatura/primera página como imagen ---
    def _cleanup_prev_tmp(self):
        try:
            if getattr(self, "_prev_tmp_file", None) and os.path.exists(self._prev_tmp_file):
                try: os.remove(self._prev_tmp_file)
                except Exception: pass
            self._prev_tmp_file = None
        except Exception:
            pass

    def _calc_canvas_size(self):
        try:
            w = int(self.prev_canvas.winfo_width())
            h = int(self.prev_canvas.winfo_height())
        except Exception:
            w = h = 0
        if w <= 1 or h <= 1:
            try:
                w = int(self.prev_canvas.winfo_reqwidth())
                h = int(self.prev_canvas.winfo_reqheight())
            except Exception:
                w, h = 600, 800
        return max(w, 1), max(h, 1)

    def _redraw_preview(self):
        """Redibuja la miniatura ajustándola al tamaño del canvas (se llama en <Configure>)."""
        if getattr(self, "_prev_pil", None) is None:
            return
        w, h = self._calc_canvas_size()
        try:
            from PIL import Image, ImageTk
            iw, ih = self._prev_pil.size
            if iw <= 0 or ih <= 0:
                return
            scale = min(w/iw, h/ih) if iw and ih else 1.0
            nw, nh = max(int(iw*scale), 1), max(int(ih*scale), 1)
            img = self._prev_pil.resize((nw, nh), Image.LANCZOS)
            tk_img = ImageTk.PhotoImage(img)
            self._prev_img = tk_img
            self.prev_canvas.delete("all")
            self.prev_canvas.create_image(w//2, h//2, image=tk_img)
        except Exception:
            # Fallback: si no hay PIL disponible, intentamos usar la última imagen Tk
            try:
                self.prev_canvas.delete("all")
                self.prev_canvas.create_image(w//2, h//2, image=self._prev_img)
            except Exception:
                pass

    def _normalize_text(self, s: str) -> str:
        """Normaliza saltos de línea:
        - Convierte \r\n / \r en \n
        - Desescapa literales "\n" y "\\r\\n" que vienen como texto plano
        - Sustituye tabs por 4 espacios
        - Compacta espacios en blanco excesivos
        """
        if not isinstance(s, str):
            try:
                s = str(s)
            except Exception:
                return ""
        # Normalizar saltos
        s = s.replace("\r\n", "\n").replace("\r", "\n")
        # Desescapar secuencias literales
        s = re.sub(r"\\r?\\n", "\n", s)
        # También arregla secuencias con barra normal ("/n", "/r/n") que a veces aparecen en texto
        s = re.sub(r"(?<!/)/(?:r?n)(?!\w)", "\n", s)
        # Tabs a espacios
        s = s.replace("\t", "    ")
        # Opcional: colapsar 3+ saltos en 2
        s = re.sub(r"\n{3,}", "\n\n", s)
        return s

    def _text_to_image(self, text, width=1000, height=1400, margin=40):
        """Renderiza texto simple a una imagen (fallback si no hay render real)."""
        text = self._normalize_text(text)
        try:
            from PIL import Image, ImageDraw, ImageFont
        except Exception:
            return None
        img = Image.new("RGB", (width, height), (255,255,255))
        draw = ImageDraw.Draw(img)
        try:
            font = ImageFont.truetype("seguiemj.ttf", 20)
        except Exception:
            try:
                font = ImageFont.truetype("arial.ttf", 20)
            except Exception:
                font = ImageFont.load_default()

        import textwrap as _tw
        max_chars = max(20, (width - 2*margin)//12)
        wrapped = []
        for line in (text or "").splitlines():
            wrapped += _tw.wrap(line, width=max_chars) or [""]
        y = margin
        for line in wrapped:
            if y > height - margin: break
            draw.text((margin, y), line, fill=(0,0,0), font=font)
            y += 26
        return img

    def _convert_word_to_pdf(self, ruta):
        """Usa MS Word (COM) para exportar a PDF. Devuelve ruta PDF temporal o None."""
        try:
            import win32com.client  # type: ignore
            import pywintypes  # type: ignore
        except Exception:
            return None
        try:
            temp_pdf = tempfile.mkstemp(suffix=".pdf")[1]
            word = win32com.client.Dispatch("Word.Application")
            word.Visible = False
            try:
                word.DisplayAlerts = 0
                word.ScreenUpdating = False
            except Exception:
                pass
            doc = word.Documents.Open(ruta, ReadOnly=True)
            doc.ExportAsFixedFormat(temp_pdf, 17)  # 17 = wdFormatPDF
            doc.Close(False)
            word.Quit()
            self._prev_tmp_file = temp_pdf
            return temp_pdf
        except Exception:
            try:
                word.Quit()
            except Exception:
                pass
            return None

    def _convert_excel_to_pdf(self, ruta):
        """Usa MS Excel (COM) para exportar a PDF. Devuelve ruta PDF temporal o None."""
        try:
            import win32com.client  # type: ignore
        except Exception:
            return None
        try:
            temp_pdf = tempfile.mkstemp(suffix=".pdf")[1]
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            wb = excel.Workbooks.Open(ruta, ReadOnly=True)
            wb.ExportAsFixedFormat(0, temp_pdf)  # 0 = xlTypePDF
            wb.Close(False)
            excel.Quit()
            self._prev_tmp_file = temp_pdf
            return temp_pdf
        except Exception:
            try:
                excel.Quit()
            except Exception:
                pass
            return None

    def _export_ppt_first_slide_png(self, ruta, width=1280, height=720):
        """Usa MS PowerPoint (COM) para exportar la primera diapositiva a PNG. Devuelve ruta PNG temporal o None."""
        try:
            import win32com.client  # type: ignore
        except Exception:
            return None
        try:
            temp_png = tempfile.mkstemp(suffix=".png")[1]
            powerpoint = win32com.client.Dispatch("PowerPoint.Application")
            powerpoint.Visible = 0
            pres = powerpoint.Presentations.Open(ruta, WithWindow=False)
            pres.Slides(1).Export(temp_png, "PNG", width, height)
            pres.Close()
            powerpoint.Quit()
            self._prev_tmp_file = temp_png
            return temp_png
        except Exception:
            try:
                powerpoint.Quit()
            except Exception:
                pass
            return None

    def _preview_first_page_to_image(self, ruta):
        """
        Intenta obtener una imagen de la primera página/diapositiva para la mayoría de formatos.
        Devuelve PIL.Image o None.
        """
        try:
            from PIL import Image
        except Exception:
            return None

        self._cleanup_prev_tmp()

        ext = os.path.splitext(ruta)[1].lower()
        media_ext = {".mp3",".wav",".flac",".aac",".ogg",".wma",".m4a",".mp4",".mov",".avi",".mkv",".wmv",".webm"}
        if ext in media_ext:
            return None

        img_ext = {".png",".jpg",".jpeg",".gif",".bmp",".tif",".tiff"}
        if ext in img_ext:
            try:
                return Image.open(ruta)
            except Exception:
                return None

        if ext == ".pdf":
            try:
                import fitz  # PyMuPDF
                doc = fitz.open(ruta)
                if len(doc):
                    page = doc[0]
                    pix = page.get_pixmap(alpha=False, dpi=120)
                    import io
                    return Image.open(io.BytesIO(pix.tobytes("png")))
            except Exception:
                try:
                    import PyPDF2
                    with open(ruta, "rb") as f:
                        reader = PyPDF2.PdfReader(f)
                        text = reader.pages[0].extract_text() if reader.pages else ""
                    return self._text_to_image(text or "Sin texto extraíble.")
                except Exception:
                    return None

        if ext in {".ppt", ".pptx"}:
            png = self._export_ppt_first_slide_png(ruta)
            if png and os.path.exists(png):
                try:
                    return Image.open(png)
                except Exception:
                    pass
            return None

        if ext in {".doc", ".docx"}:
            pdf = self._convert_word_to_pdf(ruta)
            if pdf and os.path.exists(pdf):
                try:
                    import fitz, io  # type: ignore
                    doc = fitz.open(pdf)
                    if len(doc):
                        pix = doc[0].get_pixmap(alpha=False, dpi=120)
                        return Image.open(io.BytesIO(pix.tobytes("png")))
                except Exception:
                    pass
            try:
                if ext == ".docx":
                    import zipfile, re
                    with zipfile.ZipFile(ruta) as z:
                        xml = z.read("word/document.xml").decode("utf-8", errors="ignore")
                    xml = xml.replace("</w:p>", "\n")
                    text = re.sub(r"<[^>]+>", "", xml)
                    text = text.replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">")
                    return self._text_to_image(text[:4000])
            except Exception:
                pass
            return None

        if ext in {".xls", ".xlsx", ".xlsm", ".xltx"}:
            pdf = self._convert_excel_to_pdf(ruta)
            if pdf and os.path.exists(pdf):
                try:
                    import fitz, io  # type: ignore
                    doc = fitz.open(pdf)
                    if len(doc):
                        pix = doc[0].get_pixmap(alpha=False, dpi=120)
                        return Image.open(io.BytesIO(pix.tobytes("png")))
                except Exception:
                    pass
            try:
                import openpyxl  # type: ignore
                wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
                ws = wb.active
                rows = []
                for r in ws.iter_rows(min_row=1, max_row=25, min_col=1, max_col=12, values_only=True):
                    rows.append("\\t".join("" if v is None else str(v) for v in r))
                return self._text_to_image("\n".join(rows))
            except Exception:
                try:
                    import zipfile, xml.etree.ElementTree as ET
                    z = zipfile.ZipFile(ruta)
                    sst = []
                    try:
                        root = ET.fromstring(z.read("xl/sharedStrings.xml"))
                        for si in root.iter():
                            if si.tag.endswith("t"):
                                sst.append(si.text or "")
                    except Exception:
                        pass
                    wb = ET.fromstring(z.read("xl/workbook.xml"))
                    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
                    first_sheet_id = wb.find(f".//{ns}sheet").attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
                    rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
                    target = None
                    for rel in rels:
                        if rel.attrib.get("Id")==first_sheet_id:
                            target = rel.attrib.get("Target"); break
                    sh = ET.fromstring(z.read("xl/"+target))
                    lines = []
                    for r in sh.iter():
                        if not r.tag.endswith("row"): continue
                        row_vals=[]
                        for c in r:
                            if not c.tag.endswith("c"): continue
                            t_attr = c.attrib.get("t")
                            v = c.find(f"{ns}v")
                            val = ""
                            if t_attr == "s" and v is not None:
                                try:
                                    idx = int(v.text or "0"); val = sst[idx] if 0 <= idx < len(sst) else ""
                                except Exception:
                                    val = v.text or ""
                            else:
                                val = v.text if v is not None else ""
                            row_vals.append("" if val is None else str(val))
                        if row_vals:
                            lines.append("\\t".join(row_vals[:12]))
                        if len(lines) >= 25:
                            break
                    return self._text_to_image("\n".join(lines))
                except Exception:
                    return None

        text_ext = {".txt",".py",".md",".csv",".log",".ini",".json",".xml",".yaml",".yml",".sql",".html",".htm"}
        if ext in text_ext:
            try:
                with open(ruta, "r", encoding="utf-8", errors="ignore") as f:
                    text = f.read(120000)
                return self._text_to_image(text)
            except Exception:
                return None

        return None
    def _extract_text_docx_zip(self, ruta, max_paras=60):
        import zipfile, re
        try:
            with zipfile.ZipFile(ruta) as z:
                xml = z.read("word/document.xml").decode("utf-8", errors="ignore")
        except Exception as e:
            return f"DOCX sin texto.\n{e}"
        xml = xml.replace("</w:p>", "\n")
        text = re.sub(r"<[^>]+>", "", xml)
        text = text.replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">")
        lines = [ln.strip() for ln in text.splitlines()]
        out = "\n".join([ln for ln in lines if ln][:max_paras])
        return out or "Sin texto"

    def _extract_text_xlsx_zip(self, ruta, max_rows=20, max_cols=8):
        import zipfile, xml.etree.ElementTree as ET
        try:
            z = zipfile.ZipFile(ruta)
        except Exception as e:
            return f"XLSX no legible.\n{e}"
        # shared strings
        sst = []
        try:
            sst_xml = z.read("xl/sharedStrings.xml")
            root = ET.fromstring(sst_xml)
            for si in root.iter():
                if si.tag.endswith("t"):
                    sst.append(si.text or "")
        except Exception:
            sst = []
        # first sheet path
        try:
            wb = ET.fromstring(z.read("xl/workbook.xml"))
            ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
            first_sheet_id = wb.find(f".//{ns}sheet").attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
            target = None
            for rel in rels:
                if rel.attrib.get("Id")==first_sheet_id:
                    target = rel.attrib.get("Target"); break
            sheet_xml = z.read("xl/"+target)
            sh = ET.fromstring(sheet_xml)
            rows=[]
            for r in sh.iter():
                if not r.tag.endswith("row"): continue
                row_vals=[]
                for c in r:
                    if not c.tag.endswith("c"): continue
                    t_attr = c.attrib.get("t")  # 's' (shared), 'inlineStr' o None
                    v = c.find(f"{ns}v")
                    val = ""
                    if t_attr == "s" and v is not None:
                        try:
                            idx=int(v.text or "0"); val=sst[idx] if 0 <= idx < len(sst) else ""
                        except Exception:
                            val = v.text or ""
                    elif t_attr == "inlineStr":
                        is_node = c.find(f"{ns}is")
                        texts = []
                        if is_node is not None:
                            for tnode in is_node.iter():
                                if tnode.tag.endswith("t"):
                                    texts.append(tnode.text or "")
                        val = "".join(texts)
                    else:
                        val = v.text if v is not None else ""
                    row_vals.append("" if val is None else str(val))
                if row_vals:
                    rows.append("\t".join(row_vals[:max_cols]))
                if len(rows)>=max_rows: break
            return "\n".join(rows) or "Sin datos"
        except Exception as e:
            return f"No se pudo leer hoja.\n{e}"

    def _extract_text_pptx_zip(self, ruta, max_slides=6):
        import zipfile, xml.etree.ElementTree as ET
        try:
            z = zipfile.ZipFile(ruta)
            slides = []
            i = 1
            while i<=max_slides:
                name = f"ppt/slides/slide{i}.xml"
                try:
                    xml = z.read(name)
                except Exception:
                    break
                root = ET.fromstring(xml)
                texts=[]
                for t in root.iter():
                    if t.tag.endswith("t"):
                        texts.append(t.text or "")
                slide_text = " ".join(texts).strip()
                slides.append(f"[{i}] {slide_text}")
                i += 1
            return "\n\n".join(slides) or "Sin texto"
        except Exception as e:
            return f"PPTX sin texto.\n{e}"

    def _render_preview(self, ruta):
        """
        Obtiene una imagen de la primera página/diapositiva cuando sea posible.
        Si no se puede, muestra texto como fallback.
        """
        try:
            img = self._preview_first_page_to_image(ruta)
        except Exception:
            img = None
        if img is not None:
            self.after(0, lambda: self._show_preview_image(img))
            return
        # Fallback a texto si procede
        import os
        ext = os.path.splitext(ruta)[1].lower()

        # Aviso de dependencias (una vez) si no se pudo renderizar imagen real
        try:
            if not getattr(self, "_deps_warned", False) and ext in {".doc",".docx",".ppt",".pptx",".xls",".xlsx",".pdf"}:
                self._deps_warned = True
                self.after(0, lambda: self._append_msg(
                    "Sugerencia: para miniaturas reales instala 'pywin32' (Office COM) y 'pymupdf'.", "INFO"))
        except Exception:
            pass
        text_ext = {".txt",".py",".md",".csv",".log",".ini",".json",".xml",".yaml",".yml",".sql"}
        if ext in text_ext:
            try:
                with open(ruta, "r", encoding="utf-8", errors="ignore") as f:
                    chunk = f.read(65536)
                self.after(0, lambda: self._show_preview_text(chunk))
                return
            except Exception as e:
                self.after(0, lambda: self._show_preview_text("No se pudo leer el texto.\n{}".format(e)))
                return
        self.after(0, lambda: self._show_preview_text("Sin miniatura disponible."))

    def _show_keywords_for_file(self, ruta: str):
        try:
            conn = sqlite3.connect(self._db_path(), check_same_thread=False)
            self._ensure_kw_schema(conn)
            cur = conn.cursor()
            rows = cur.execute("SELECT keyword, freq FROM doc_keywords WHERE fullpath = ? ORDER BY keyword", (ruta,)).fetchall()
            conn.close()
        except Exception as e:
            self._append_msg(f"Error leyendo keywords: {e}", "WARN")
            rows = []
        if not rows:
            messagebox.showinfo(APP_NAME, "No hay palabras clave guardadas para este fichero.")
            return
        txt = "\n".join([f"- {k} (freq {f})" for k,f in rows])
        messagebox.showinfo(APP_NAME, f"Palabras clave:\n\n{txt}")

    def _open_help(self, event=None):
        HelpDialog(self.master, APP_NAME, APP_VERSION)



    


    def _ensure_kw_schema(self, conn):
        try:
            c = conn.cursor()
            c.execute("""                CREATE TABLE IF NOT EXISTS doc_keywords(
                    id INTEGER PRIMARY KEY,
                    fullpath TEXT NOT NULL,
                    name TEXT,
                    ext TEXT,
                    keyword TEXT NOT NULL,
                    freq INTEGER DEFAULT 1,
                    created_at TEXT DEFAULT (datetime('now'))
                )
            """ )
            c.execute("CREATE INDEX IF NOT EXISTS idx_kw_path ON doc_keywords(fullpath)")
            c.execute("CREATE INDEX IF NOT EXISTS idx_kw_kw ON doc_keywords(keyword)")
            conn.commit()
        except Exception as e:
            self._append_msg(f"SQLite (keywords) error: {e}", "WARN")

    def _extract_text_generic(self, ruta: str, max_chars: int = 250_000) -> str:
        ruta = str(ruta)
        ext = os.path.splitext(ruta)[1].lower()
        try:
            if ext in {'.txt','.py','.md','.csv','.log','.ini','.json','.xml','.yaml','.yml','.sql','.html','.htm'}:
                with open(ruta, 'r', encoding='utf-8', errors='ignore') as f:
                    return f.read(max_chars)
            if ext == '.docx':
                return self._extract_text_docx_zip(ruta)
            if ext == '.pptx':
                return self._extract_text_pptx_zip(ruta)
            if ext in {'.xlsx','.xlsm','.xltx'}:
                return self._extract_text_xlsx_zip(ruta)
            if ext == '.pdf':
                try:
                    import fitz
                    doc = fitz.open(ruta)
                    texts = []
                    for i, page in enumerate(doc):
                        if i >= 8: break
                        t = page.get_text()
                        if t: texts.append(t)
                    return "\n".join(texts)[:max_chars] if texts else ""
                except Exception:
                    try:
                        import PyPDF2
                        with open(ruta, 'rb') as f:
                            reader = PyPDF2.PdfReader(f)
                            texts = []
                            for i, p in enumerate(reader.pages):
                                if i >= 8: break
                                try:
                                    texts.append(p.extract_text() or "")
                                except Exception:
                                    pass
                            return "\n".join(texts)[:max_chars]
                    except Exception:
                        return ""
            if ext in {'.doc','.xls','.ppt'}:
                return ""
        except Exception as e:
            self._append_msg(f"Error extrayendo texto de {ruta}: {e}", "WARN")
        return ""

    def _extract_noun_frequencies(self, text: str, use_spacy: bool = True) -> dict:
        freqs = collections.Counter()
        if not text or len(text) < 10:
            return freqs
        if use_spacy:
            try:
                import spacy
                nlp = None
                try:
                    nlp = spacy.load("es_core_news_sm")
                except Exception:
                    try:
                        nlp = spacy.load("es_core_news_md")
                    except Exception:
                        nlp = None
                if nlp is not None:
                    doc = nlp(text[:200_000])
                    for tok in doc:
                        if tok.pos_ in ("NOUN","PROPN"):
                            w = tok.lemma_.lower().strip()
                            if w and w not in SPANISH_STOPWORDS and len(w) >= 3 and not w.endswith("mente"):
                                freqs[w] += 1
                    return dict(freqs)
            except Exception:
                pass
        tokens = _simple_tokenize_es(text)
        for t in tokens:
            if _is_probable_noun_es(t):
                freqs[t] += 1
        return dict(freqs)

    def cmd_scraper(self):
        targets = []
        sel = self.tree.selection()
        if sel:
            for iid in sel:
                vals = self.tree.item(iid, "values")
                if vals and len(vals) >= 6:
                    targets.append(vals[5])
        else:
            for iid in self.tree.get_children():
                vals = self.tree.item(iid, "values")
                if vals and len(vals) >= 6:
                    targets.append(vals[5])
        if not targets:
            messagebox.showinfo(APP_NAME, "No hay ficheros seleccionados ni visibles para analizar.")
            return
        if len(targets) > 1:
            self._append_msg(f"Scraper: {len(targets)} archivos detectados; procesando el primero ahora (versión inicial).", "INFO")
        ruta = targets[0]
        nombre = os.path.basename(ruta)
        ext = os.path.splitext(nombre)[1].lstrip('.').lower()
        self._task_open("Extrayendo y analizando…", total=3)
        self._task_update(1, "Extrayendo texto…")
        text = self._extract_text_generic(ruta)
        self._task_update(2, "Determinando sustantivos…")
        freqs = self._extract_noun_frequencies(text)
        freqs = {k:v for k,v in freqs.items() if v >= 2}
        if not freqs:
            self._task_close()
            messagebox.showinfo(APP_NAME, "No se han encontrado sustantivos repetidos (umbral ≥2).")
            return
        self._task_update(3, "Mostrando selección…")
        self._task_close()
        self._open_keywords_dialog(ruta, nombre, ext, freqs)

    def _open_keywords_dialog(self, fullpath: str, nombre: str, ext: str, freqs: dict):
        dlg = tk.Toplevel(self)
        dlg.title(f"Scraper de palabras clave – {os.path.basename(fullpath)}")
        dlg.minsize(560, 520)
        dlg.resizable(True, True)
        # dlg.transient(self.winfo_toplevel())  # desactivado para permitir min/max
        dlg.grab_set()
        frm_top = ttk.Frame(dlg, padding=(10,8)); frm_top.pack(fill="x")
        ttk.Label(frm_top, text=os.path.basename(fullpath)).pack(side="left")
        frm_ctrl = ttk.Frame(dlg, padding=(10,2)); frm_ctrl.pack(fill="x")
        var_umbral = tk.IntVar(value=2)
        ttk.Label(frm_ctrl, text="Umbral (freq ≥)").pack(side="left")
        spn = ttk.Spinbox(frm_ctrl, from_=1, to=10, width=4, textvariable=var_umbral)
        spn.pack(side="left", padx=(6,12))
        ttk.Separator(dlg, orient="horizontal").pack(fill="x", pady=(4,4))
        frm_add = ttk.Frame(dlg, padding=(10,2)); frm_add.pack(fill="x")
        ttk.Label(frm_add, text="Añadir palabra clave:").pack(side="left")
        var_add = tk.StringVar()
        ent_add = ttk.Entry(frm_add, width=28, textvariable=var_add); ent_add.pack(side="left", padx=(6,6))
        ttk.Button(frm_add, text="Añadir", command=lambda: add_word()).pack(side="left")
        ttk.Separator(dlg, orient="horizontal").pack(fill="x", pady=(4,4))
        # Ordenación
        var_order = tk.StringVar(value="freq")  # "freq" (desc) | "alpha" (asc)
        ttk.Label(frm_ctrl, text="Orden:").pack(side="left", padx=(12,4))
        ttk.Radiobutton(frm_ctrl, text="Frecuencia", value="freq", variable=var_order).pack(side="left")
        ttk.Radiobutton(frm_ctrl, text="Alfabético", value="alpha", variable=var_order).pack(side="left", padx=(6,0))
        ttk.Button(frm_ctrl, text="Maximizar", command=lambda: dlg.state("zoomed")).pack(side="right")
        ttk.Button(frm_ctrl, text="Minimizar", command=dlg.iconify).pack(side="right", padx=(0,6))

        ttk.Separator(dlg, orient="horizontal").pack(fill="x", pady=(4,4))
        frm_list = ttk.Frame(dlg); frm_list.pack(fill="both", expand=True)
        canvas = tk.Canvas(frm_list, bd=0, highlightthickness=0)
        ysb = ttk.Scrollbar(frm_list, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=ysb.set)
        ysb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        inner = ttk.Frame(canvas)
        winid = canvas.create_window((0,0), window=inner, anchor="nw")

        # Vars por palabra para preservar selección entre renderizados
        vars = {}
        widgets = {}

        def get_sorted_items():
            if var_order.get() == "alpha":
                return sorted(freqs.items(), key=lambda kv: (kv[0], -kv[1]))
            return sorted(freqs.items(), key=lambda kv: (-kv[1], kv[0]))

        def render_items():
            # Limpiar
            for child in list(inner.children.values()):
                child.destroy()
            # Pintar en el orden elegido
            for i, (w, f) in enumerate(get_sorted_items()):
                if w not in vars:
                    vars[w] = tk.IntVar(value=1 if f >= var_umbral.get() else 0)
                cb = ttk.Checkbutton(inner, text=f"{w} ({f})", variable=vars[w])
                cb.grid(row=i, column=0, sticky="w", padx=(6,0), pady=2)
                widgets[w] = cb
            canvas.update_idletasks()
            canvas.configure(scrollregion=canvas.bbox("all"))

        def on_configure(event):
            # Estirar contenido al ancho disponible
            canvas.itemconfig(winid, width=event.width)
            canvas.configure(scrollregion=canvas.bbox("all"))

        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", on_configure)

        # Botonera inferior
        ttk.Separator(dlg, orient="horizontal").pack(fill="x", pady=(4,4))
        frm_btns = ttk.Frame(dlg, padding=(10,8)); frm_btns.pack(fill="x")
        def select_all(v=1):
            for k in vars: vars[k].set(v)
        ttk.Button(frm_btns, text="Todos", command=lambda: select_all(1)).pack(side="left")
        ttk.Button(frm_btns, text="Ninguno", command=lambda: select_all(0)).pack(side="left", padx=(6,0))
        status = ttk.Label(frm_btns, text=""); status.pack(side="left", padx=(14,0))

        def apply_selection():
            umbral = var_umbral.get()
            selected = 0
            for (w, f) in get_sorted_items():
                want = 1 if f >= umbral else 0
                if w not in vars:
                    vars[w] = tk.IntVar(value=want)
                else:
                    vars[w].set(want)
                selected += want
            status.config(text=f"Seleccionadas: {selected}")

        ttk.Button(frm_ctrl, text="Aplicar selección", command=apply_selection).pack(side="left", padx=(0,6))

        def add_word():
            w = var_add.get().strip().lower()
            if not w:
                return
            # si no existe, crea con freq 1
            if w not in freqs:
                freqs[w] = 1
            vars.setdefault(w, tk.IntVar(value=1)).set(1)
            var_add.set("")
            render_items()

        def guardar():
            kw = [w for w,v in vars.items() if v.get()==1]
            self._save_keywords(fullpath, nombre, ext, kw, freqs)
            self._append_msg(f"Palabras clave guardadas para {nombre}: {', '.join(kw)}", "OK")
            messagebox.showinfo(APP_NAME, "Guardado con éxito")
            dlg.destroy()

        ttk.Button(frm_btns, text="Guardar", command=guardar).pack(side="right")
        ttk.Button(frm_btns, text="Cancelar", command=dlg.destroy).pack(side="right", padx=(6,0))

        # Re-render al cambiar orden
        var_order.trace_add("write", lambda *args: render_items())

        # Render inicial
        render_items()

    def _save_keywords(self, fullpath: str, nombre: str, ext: str, keywords: list[str], freqs: dict):
        try:
            conn = sqlite3.connect(self._db_path(), check_same_thread=False)
            self._ensure_kw_schema(conn)
            cur = conn.cursor()
            cur.execute("DELETE FROM doc_keywords WHERE fullpath = ?", (fullpath,))
            for k in keywords:
                cur.execute("INSERT INTO doc_keywords(fullpath, name, ext, keyword, freq) VALUES (?,?,?,?,?)",
                            (fullpath, nombre, ext, k, int(freqs.get(k, 1))))
            conn.commit()
            conn.close()
        except Exception as e:
            self._append_msg(f"Error guardando keywords: {e}", "WARN")

    # ============================ EXPORTACIÓN RÁPIDA (STREAMING, THREAD) ============================
    def cmd_exportar_excel_rapido(self):
        """Exporta TODO el índice usando openpyxl en modo write_only y en un hilo de fondo.
        Ideal para 100k–200k archivos."""
        if not self.file_index:
            messagebox.showinfo(APP_NAME, "No hay índice cargado. Pulsa ESCANEAR primero.")
            return
        default_name = f"Listado_DocuSICOP_TODO_RAPIDO_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        save_path = filedialog.asksaveasfilename(
            title="Guardar Excel (rápido)", defaultextension=".xlsx",
            initialfile=default_name, filetypes=[("Excel", "*.xlsx")]
        )
        if not save_path:
            return
        threading.Thread(target=self._worker_exportar_excel_stream, args=(save_path,), daemon=True).start()

    def _worker_exportar_excel_stream(self, save_path: str):
        try:
            total = len(self.file_index)
            self.queue.put(("task_open", ("Generando Excel (rápido)", total)))

            rutas = [Path(e["ruta"]) for e in self.file_index]
            base = self.base_path if self.base_path else self._lca_base(rutas)
            base_name = base.name

            # Calcular nº de subcarpetas máximo para columnas SUBCARPETA i
            max_sublevels = 0
            for ruta in rutas:
                dirs = self._split_dirs(base, ruta)
                sublevels = max(0, len(dirs) - 1) if dirs else 0
                if sublevels > max_sublevels:
                    max_sublevels = sublevels

            sub_headers = [f"SUBCARPETA {i}" for i in range(1, max_sublevels + 1)]
            headers = (['CARPETA BASE', 'CARPETA'] + sub_headers +
                       ['FICHERO', 'VERSIÓN', 'FECHA', 'CÓDIGO', 'CÓDIGO ANTERIOR',
                        'LOCALIZACIÓN (pendiente de actualizar)', 'FIRMAS', 'RENOVACIÓN',
                        'CAMBIOS EN 2008', 'CAMBIOS EN 2009/2010', 'CAMBIOS EN 2011', 'CAMBIOS EN 2012',
                        'CAMBIOS EN 2013', 'CAMBIOS EN 2014', 'CAMBIOS EN 2015', 'CAMBIOS EN 2016', 'CAMBIOS EN 2017',
                        'CAMBIOS EN 2018', 'CAMBIOS EN 2019', 'CAMBIOS EN 2020', 'CAMBIOS EN 2021', 'CAMBIOS EN 2022',
                        'CAMBIOS EN 2023', 'CAMBIOS EN 2024', 'ÚLTIMOS CAMBIOS (2025)', 'PENDIENTE'])
            header_to_idx = {h: i for i, h in enumerate(headers)}

            wb = Workbook(write_only=True)
            ws = wb.active
            ws.title = "Listado de documentación"

            head_fill = PatternFill("solid", fgColor="E2EFDA")
            head_font = Font(bold=True)
            header_row = []
            for h in headers:
                c = WriteOnlyCell(ws, value=h)
                c.fill = head_fill
                c.font = head_font
                c.alignment = Alignment(wrap_text=True, vertical="center")
                header_row.append(c)
            ws.append(header_row)

            pushed = 0
            for e in self.file_index:
                ruta = Path(e["ruta"])
                try:
                    fecha_dt = datetime.fromtimestamp(e.get("mod_ts", 0))
                    fecha_val = fecha_dt.strftime("%Y-%m-%d %H:%M")
                except Exception:
                    fecha_dt = None
                    fecha_val = ""

                dirs = self._split_dirs(base, ruta)
                carpeta = dirs[0] if dirs else ""
                subs = (dirs[1:] if len(dirs) > 1 else [])
                subs = subs + [""] * (max_sublevels - len(subs))

                version = self._guess_version(ruta.stem)
                codigo = self._guess_code(ruta.stem)

                try:
                    rel_parent = ruta.parent.relative_to(base)
                    rel_str = "." if str(rel_parent) == "." else str(rel_parent).replace("/", "\\")
                except Exception:
                    rel_str = ""
                localizacion = base_name if rel_str in ("", ".") else f"{base_name}\\{rel_str}"

                row_vals = [""] * len(headers)
                row_vals[header_to_idx['CARPETA BASE']] = base_name
                row_vals[header_to_idx['CARPETA']] = carpeta
                for i, sh in enumerate(sub_headers):
                    row_vals[header_to_idx[sh]] = subs[i] if i < len(subs) else ""

                cfile = WriteOnlyCell(ws, value=ruta.name)
                try:
                    cfile.hyperlink = str(ruta)
                    cfile.style = "Hyperlink"
                except Exception:
                    pass
                row_vals[header_to_idx['FICHERO']] = cfile
                row_vals[header_to_idx['VERSIÓN']] = version
                row_vals[header_to_idx['FECHA']] = fecha_val
                row_vals[header_to_idx['CÓDIGO']] = codigo
                row_vals[header_to_idx['CÓDIGO ANTERIOR']] = ""
                row_vals[header_to_idx['LOCALIZACIÓN (pendiente de actualizar)']] = localizacion
                row_vals[header_to_idx['FIRMAS']] = ""
                row_vals[header_to_idx['RENOVACIÓN']] = ""

                if fecha_dt is not None:
                    msg = f"Modificado {fecha_dt.strftime('%Y-%m-%d %H:%M')}"
                    idx = self._col_for_year(fecha_dt.year, headers, header_to_idx)
                    row_vals[idx] = msg

                ws.append(row_vals)
                pushed += 1
                if pushed % 500 == 0:
                    self.queue.put(("task_update", (500, f"{pushed} / {total}")))

            self.queue.put(("task_indet", ("Guardando archivo…",)))
            wb.save(save_path)
            self.queue.put(("task_close", None))
            self.queue.put(("msg", (f"Excel guardado en: {save_path}", "OK")))
        except Exception as e:
            self.queue.put(("task_close", None))
            self.queue.put(("msg", (f"ERROR exportando (rápido): {e}", "ERR")))
    def cmd_exportar_excel_rapido_visibles(self):
        """Exporta SOLO las filas visibles de la tabla (rápido)."""
        tv = self.tree
        if not tv.get_children():
            messagebox.showinfo(APP_NAME, "No hay filas visibles para exportar.")
            return
        from datetime import datetime

        default_name = f"Listado_DocuSICOP_VISIBLES_RAPIDO_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        save_path = filedialog.asksaveasfilename(
            title="Guardar Excel (visibles, rápido)", defaultextension=".xlsx",
            initialfile=default_name, filetypes=[("Excel", "*.xlsx")]
        )
        if not save_path:
            return

        rutas = []
        rows = []
        for iid in tv.get_children():
            nombre, ext, tam_str, mod_str, carpeta, ruta = tv.item(iid, "values")
            rutas.append(Path(ruta))
            try:
                dt = datetime.strptime(mod_str, "%Y-%m-%d %H:%M")
            except Exception:
                dt = None
            rows.append((nombre, ext, tam_str, dt, carpeta, ruta))

        base = self.base_path if self.base_path else self._lca_base(rutas)
        base_name = base.name

        max_sublevels = 0
        for ruta in rutas:
            dirs = self._split_dirs(base, ruta)
            sublevels = max(0, len(dirs) - 1) if dirs else 0
            if sublevels > max_sublevels:
                max_sublevels = sublevels

        sub_headers = [f"SUBCARPETA {i}" for i in range(1, max_sublevels + 1)]
        headers = (['CARPETA BASE', 'CARPETA'] + sub_headers +
                   ['FICHERO', 'VERSIÓN', 'FECHA', 'CÓDIGO', 'CÓDIGO ANTERIOR',
                    'LOCALIZACIÓN (pendiente de actualizar)', 'FIRMAS', 'RENOVACIÓN',
                    'CAMBIOS EN 2008', 'CAMBIOS EN 2009/2010', 'CAMBIOS EN 2011', 'CAMBIOS EN 2012',
                    'CAMBIOS EN 2013', 'CAMBIOS EN 2014', 'CAMBIOS EN 2015', 'CAMBIOS EN 2016', 'CAMBIOS EN 2017',
                    'CAMBIOS EN 2018', 'CAMBIOS EN 2019', 'CAMBIOS EN 2020', 'CAMBIOS EN 2021', 'CAMBIOS EN 2022',
                    'CAMBIOS EN 2023', 'CAMBIOS EN 2024', 'ÚLTIMOS CAMBIOS (2025)', 'PENDIENTE'])
        header_to_idx = {h: i for i, h in enumerate(headers)}

        wb = Workbook(write_only=True)
        ws = wb.active
        ws.title = "Listado de documentación"

        head_fill = PatternFill("solid", fgColor="E2EFDA")
        head_font = Font(bold=True)
        header_row = []
        for h in headers:
            c = WriteOnlyCell(ws, value=h)
            c.fill = head_fill
            c.font = head_font
            c.alignment = Alignment(wrap_text=True, vertical="center")
            header_row.append(c)
        ws.append(header_row)

        for (nombre, ext, tam_str, dt, carpeta, ruta_str) in rows:
            ruta = Path(ruta_str)
            dirs = self._split_dirs(base, ruta)
            carpeta0 = dirs[0] if dirs else ""
            subs = (dirs[1:] if len(dirs) > 1 else [])
            subs = subs + [""] * (max_sublevels - len(subs))

            version = self._guess_version(ruta.stem)
            codigo = self._guess_code(ruta.stem)

            try:
                rel_parent = ruta.parent.relative_to(base)
                rel_str = "." if str(rel_parent) == "." else str(rel_parent).replace("/", "\\")
            except Exception:
                rel_str = ""
            localizacion = base_name if rel_str in ("", ".") else f"{base_name}\\{rel_str}"

            row_vals = [""] * len(headers)
            row_vals[header_to_idx['CARPETA BASE']] = base_name
            row_vals[header_to_idx['CARPETA']] = carpeta0
            for i, sh in enumerate(sub_headers):
                row_vals[header_to_idx[sh]] = subs[i] if i < len(subs) else ""

            cfile = WriteOnlyCell(ws, value=nombre)
            try:
                cfile.hyperlink = str(ruta)
                cfile.style = "Hyperlink"
            except Exception:
                pass
            row_vals[header_to_idx['FICHERO']] = cfile
            row_vals[header_to_idx['VERSIÓN']] = version
            row_vals[header_to_idx['FECHA']] = dt.strftime("%Y-%m-%d %H:%M") if dt else ""
            row_vals[header_to_idx['CÓDIGO']] = codigo
            row_vals[header_to_idx['CÓDIGO ANTERIOR']] = ""
            row_vals[header_to_idx['LOCALIZACIÓN (pendiente de actualizar)']] = localizacion
            row_vals[header_to_idx['FIRMAS']] = ""
            row_vals[header_to_idx['RENOVACIÓN']] = ""

            if dt is not None:
                msg = f"Modificado {dt.strftime('%Y-%m-%d %H:%M')}"
                idx = self._col_for_year(dt.year, headers, header_to_idx)
                row_vals[idx] = msg

            ws.append(row_vals)

        wb.save(save_path)
        self._append_msg(f"Excel guardado en: {save_path}", "OK")
        messagebox.showinfo(APP_NAME, f"Excel generado:\n{save_path}")

class HelpDialog(tk.Toplevel):
    def __init__(self, master, app_name: str, version: str):
        super().__init__(master)
        self.title(f"{app_name} • Ayuda")
        self.transient(master); self.resizable(True, True); self.grab_set()
        self.minsize(780, 560)

        outer = ttk.Frame(self, padding=10); outer.pack(fill="both", expand=True)

        base = tkfont.nametofont("TkDefaultFont").copy()
        try: base.configure(family="Segoe UI", size=10)
        except Exception: base.configure(size=10)
        self.f_base = base
        self.f_bold = base.copy(); self.f_bold.configure(weight="bold")
        self.f_italic = base.copy(); self.f_italic.configure(slant="italic")
        self.f_code = tkfont.Font(family="Consolas", size=10)
        self.f_h1 = base.copy(); self.f_h1.configure(size=14, weight="bold")
        self.f_h2 = base.copy(); self.f_h2.configure(size=12, weight="bold")
        self.f_h3 = base.copy(); self.f_h3.configure(size=11, weight="bold")

        # Barra filtro
        bar = ttk.Frame(outer); bar.pack(fill="x", pady=(0,8))
        ttk.Label(bar, text="Filtrar en esta pestaña:").pack(side="left")
        self.var_filter = tk.StringVar()
        ent = ttk.Entry(bar, textvariable=self.var_filter, width=40)
        ent.pack(side="left", padx=(8,8)); ent.bind("<Return>", self._on_search_enter)
        ttk.Button(bar, text="Buscar", command=self._on_search).pack(side="left")
        ttk.Button(bar, text="Anterior", command=self._on_prev).pack(side="left", padx=(6,0))
        ttk.Button(bar, text="Siguiente", command=self._on_next).pack(side="left", padx=(6,8))
        ttk.Button(bar, text="Limpiar", command=self._on_clear).pack(side="left")
        self.lbl_status = ttk.Label(bar, text=""); self.lbl_status.pack(side="right")

        # Notebook + tags
        self.nb = ttk.Notebook(outer); self.nb.pack(fill="both", expand=True)
        self.nb.bind("<<NotebookTabChanged>>", self._on_tab_changed)
        self._text_by_tab = {}; self._matches_by_tab = {}; self._cursor_by_tab = {}
        self._indent_tags = set()

        def add_tab(title: str, md_text: str):
            frame = ttk.Frame(self.nb); self.nb.add(frame, text=title)
            txt = ScrolledText(frame, wrap="word"); txt.pack(fill="both", expand=True)
            txt.configure(font=self.f_base)
            txt.tag_configure("match", background="#fff59d")
            txt.tag_configure("current", background="#ffe082", underline=True)
            txt.tag_configure("bold", font=self.f_bold)
            txt.tag_configure("italic", font=self.f_italic)
            txt.tag_configure("code", font=self.f_code, background="#f4f4f5")
            txt.tag_configure("h1", font=self.f_h1, spacing1=6, spacing3=4)
            txt.tag_configure("h2", font=self.f_h2, spacing1=4, spacing3=2)
            txt.tag_configure("h3", font=self.f_h3)
            txt.tag_configure("bullet", lmargin1=20, lmargin2=36)
            self._insert_md(txt, md_text); txt.configure(state="disabled")
            self._text_by_tab[frame]=txt; self._matches_by_tab[frame]=[]; self._cursor_by_tab[frame]=-1
            return frame

        guia = r"""
        # Guía rápida

        1) **Selecciona Carpeta Base** (pestaña *Base*).
        2) Pulsa **Escanear (F5)** para indexar toda la jerarquía en SQLite.
        3) Escribe texto y pulsa **Enter** o **Buscar**.
           - Puedes marcar **"Buscar también en ruta"** para incluir directorios.
           - Filtra por **Extensiones (csv)**, p. ej.: `pdf,docx,xlsx`.
        4) Revisa la pestaña **Resultados**. Clic derecho sobre un resultado:
           **Abrir fichero** o **Abrir carpeta**.
        5) **Exportar ▾**: elige *Visibles* o *Todo* para crear Excel con:
           - **CARPETA BASE** (columna independiente)
           - **LOCALIZACIÓN relativa** `"<BASE>\\sub\\sub"`
           - **Hipervínculo** absoluto (Windows/UNC).
        """.strip("\n")
        add_tab("Guía rápida", guia)

        atajos = r"""
        # Atajos
        - **F1**: abrir ayuda.
        - **F5**: escanear carpeta base.
        - **Enter**: ejecutar búsqueda.
        - **Ctrl+L**: limpiar resultados/mensajes.
        - **Doble clic** en resultado: abrir fichero.
        - **Botón derecho**: menú contextual.
        """.strip("\n")
        add_tab("Atajos", atajos)

        busqueda = r"""
        # Búsqueda
        - Separa términos por espacios (se aplican todos).
        - **Comodines**: `*` o `%` (cualquier secuencia), `?` o `_` (un carácter).
        - **Escapar** un comodín literal con `\\`: `\\*`, `\\?`, `\\%`, `\\_`.
        - **Buscar también en ruta**: incluye coincidencias en la carpeta/ruta completa.
        - **Extensiones (csv)**: lista sin puntos, p. ej. `pdf,docx,xlsx`.
        """.strip("\n")
        add_tab("Búsqueda", busqueda)

        previ = r"""
        # Previsualización
        - Si está `pymupdf`, se genera miniatura real para **PDF**.
        - Con **Microsoft Office** + `pywin32`, se generan miniaturas para DOCX/XLSX/PPTX.
        - Para **texto** (txt/csv/md/json/sql…) se muestra un visor de texto.
        """.strip("\n")
        add_tab("Previsualización", previ)

        export = r"""
        # Exportación
        - **Visibles**: exporta lo que ves en la tabla.
        - **Todo**: exporta todo el índice.
        - Columnas: **CARPETA BASE**, **LOCALIZACIÓN** (`<BASE>\\sub\\sub`), hipervínculo absoluto, etc.
        """.strip("\n")
        add_tab("Exportación", export)

        problemas = r"""
        # Problemas frecuentes
        - **Migraciones SQLite**: si ves mensajes de columnas faltantes, borra `index_cache.sqlite` y vuelve a escanear.
        - **Sin miniaturas PDF/Office**: instala `pymupdf` / `pywin32`.
        - **Rutas UNC**: `\\\\servidor\\recurso\\carpeta`.
        """.strip("\n")
        add_tab("Problemas", problemas)

        acerca = f"""
        # Acerca de
        **{app_name}** v{version}
        Utilidad para indexar carpetas, buscar con comodines y exportar a Excel.
        """.strip("\n")
        add_tab("Acerca de", acerca)

        btns = ttk.Frame(outer); btns.pack(fill="x", pady=(10,0))
        ttk.Button(btns, text="Cerrar", command=self.destroy).pack(side="right")
        ent.focus_set(); self._update_status()

    def _ensure_indent_tag(self, txt, base_name: str, level: int):
        tag = f"{base_name}{level}"
        try: txt.tag_cget(tag, "lmargin1")
        except Exception: txt.tag_configure(tag, lmargin1=20+level*20, lmargin2=36+level*20)
        return tag

    def _insert_md(self, txt: ScrolledText, md: str):
        import textwrap as _tw
        md = _tw.dedent(md.strip("\n")); lines = md.splitlines()
        txt.configure(state="normal")
        for line in lines:
            raw=line.rstrip("\n"); stripped=raw.lstrip(); indent=len(raw)-len(stripped); level=max(indent//2,0)
            if stripped.startswith("# "):  self._insert_inline(txt, stripped[2:].strip(), ("h1",)); txt.insert("end","\n"); continue
            if stripped.startswith("## "): self._insert_inline(txt, stripped[3:].strip(), ("h2",)); txt.insert("end","\n"); continue
            if stripped.startswith("### "):self._insert_inline(txt, stripped[4:].strip(), ("h3",)); txt.insert("end","\n"); continue
            if stripped.startswith(("- ","• ")):
                tag=self._ensure_indent_tag(txt,"bullet",level); self._insert_inline(txt, "• "+stripped[2:].strip(), (tag,)); txt.insert("end","\n"); continue
            m = re.match(r"^(\d+\))\s+(.*)$", stripped)
            if m: 
                num,rest=m.groups(); tag=self._ensure_indent_tag(txt,"num",level)
                txt.insert("end", num+" ", (tag,"bold")); self._insert_inline(txt, rest, (tag,)); txt.insert("end","\n"); continue
            if stripped=="": txt.insert("end","\n")
            else: self._insert_inline(txt, stripped); txt.insert("end","\n")
        txt.configure(state="disabled")

    def _insert_inline(self, txt: ScrolledText, s: str, base_tags=()):
        i=0; N=len(s)
        while i<N:
            if i+1<N and s[i:i+2]=="**":
                j=s.find("**", i+2)
                if j!=-1: txt.insert("end", s[i+2:j], (*base_tags,"bold")); i=j+2; continue
            if s[i]=="`":
                j=s.find("`", i+1)
                if j!=-1: txt.insert("end", s[i+1:j], (*base_tags,"code")); i=j+1; continue
            if s[i]=="*":
                j=s.find("*", i+1)
                if j!=-1: txt.insert("end", s[i+1:j], (*base_tags,"italic")); i=j+1; continue
            txt.insert("end", s[i], base_tags); i+=1

    # Búsqueda
    def _current_frame(self): return self.nb.nametowidget(self.nb.select())
    def _clear_highlight(self, txt):
        st=str(txt.cget("state")); txt.configure(state="normal")
        txt.tag_remove("match","1.0","end"); txt.tag_remove("current","1.0","end"); txt.configure(state=st)
    def _highlight(self, txt, pattern):
        self._clear_highlight(txt); 
        if not pattern: return []
        st=str(txt.cget("state")); txt.configure(state="normal")
        start="1.0"; matches=[]
        while True:
            pos=txt.search(pattern, start, stopindex="end", nocase=True)
            if not pos: break
            end=txt.index(f"{pos}+{len(pattern)}c"); txt.tag_add("match", pos, end)
            matches.append((pos,end)); start=end
        txt.configure(state=st); return matches
    def _apply_search_current(self):
        pattern=self.var_filter.get().strip(); frm=self._current_frame(); txt=self._text_by_tab.get(frm)
        if not txt: return
        matches=self._highlight(txt, pattern); self._matches_by_tab[frm]=matches; self._cursor_by_tab[frm]=(0 if matches else -1)
        self._focus_current(); self._update_status()
    def _focus_current(self):
        frm=self._current_frame(); txt=self._text_by_tab.get(frm); matches=self._matches_by_tab.get(frm, []); cur=self._cursor_by_tab.get(frm, -1)
        if txt:
            st=str(txt.cget("state")); txt.configure(state="normal"); txt.tag_remove("current","1.0","end")
            if 0<=cur<len(matches): s,e=matches[cur]; txt.tag_add("current", s, e); txt.see(s)
            txt.configure(state=st)
    def _update_status(self):
        frm=self._current_frame(); total=len(self._matches_by_tab.get(frm, [])); cur=self._cursor_by_tab.get(frm, -1)
        self.lbl_status.config(text=("0 coincidencias" if total==0 else f"Coincidencias: {total}  •  Posición: {cur+1}/{total}"))
    def _on_search_enter(self, event): self._apply_search_current()
    def _on_search(self): self._apply_search_current()
    def _on_next(self):
        frm=self._current_frame(); total=len(self._matches_by_tab.get(frm, []))
        if total==0: return self._update_status()
        self._cursor_by_tab[frm]=(self._cursor_by_tab[frm]+1)%total; self._focus_current(); self._update_status()
    def _on_prev(self):
        frm=self._current_frame(); total=len(self._matches_by_tab.get(frm, []))
        if total==0: return self._update_status()
        self._cursor_by_tab[frm]=(self._cursor_by_tab[frm]-1)%total; self._focus_current(); self._update_status()
    def _on_clear(self):
        self.var_filter.set(""); frm=self._current_frame(); txt=self._text_by_tab.get(frm)
        if txt: self._clear_highlight(txt); self._matches_by_tab[frm]=[]; self._cursor_by_tab[frm]=-1; self._update_status()
    def _on_tab_changed(self, event): self._apply_search_current()


    # ============================ RAG / EMBEDDINGS (base-aware) ============================
    def _db_rag_ensure(self, conn):
        try:
            c = conn.cursor()
            c.execute("CREATE TABLE IF NOT EXISTS chunks(id INTEGER PRIMARY KEY, file_path TEXT, mtime REAL, text TEXT)")
            c.execute("CREATE TABLE IF NOT EXISTS embeddings(chunk_id INTEGER PRIMARY KEY, vec BLOB)")
            c.execute("CREATE INDEX IF NOT EXISTS idx_chunks_path ON chunks(file_path)")
            c.execute("CREATE INDEX IF NOT EXISTS idx_chunks_mtime ON chunks(mtime)")
            conn.commit()
        except Exception as e:
            self._append_msg(f"SQLite (RAG) error: {e}", "WARN")

    def _vec_to_blob(self, v):
        try:
            return struct.pack('<'+'f'*len(v), *v)
        except Exception:
            import array
            return array.array('f', v).tobytes()

    def _blob_to_vec(self, b: bytes):
        n = len(b)//4
        return list(struct.unpack('<'+'f'*n, b)) if n>0 else []

    def _hash_embedder(self, text: str, dim: int = 256):
        v = [0.0]*dim
        if not text:
            return v
        for tok in re.findall(r"[A-Za-zÁÉÍÓÚÜáéíóúüÑñ0-9]{2,}", text.lower()):
            if tok in SPANISH_STOPWORDS:
                continue
            h = int(hashlib.md5(tok.encode('utf-8')).hexdigest(), 16)
            i = h % dim
            s = 1.0 if (h >> 1) & 1 else -1.0
            v[i] += s
        import math as _m
        norm = _m.sqrt(sum(x*x for x in v)) or 1.0
        return [x/norm for x in v]

    def _text_chunks(self, txt: str, max_chars=1200, overlap=200):
        txt = (txt or "").replace("\r\n","\n").replace("\r","\n")
        parts = []
        step = max(1, max_chars - overlap)
        i = 0; N = len(txt)
        while i < N:
            parts.append(txt[i:i+max_chars])
            i += step
        return parts

    def _index_file_chunks(self, fullpath: str, mtime_ts: float):
        try:
            conn = getattr(self, "_db_conn", None)
            if conn is None:
                conn = sqlite3.connect(self._db_path(), check_same_thread=False)
            self._db_rag_ensure(conn)
            c = conn.cursor()
            try:
                c.execute("DELETE FROM embeddings WHERE chunk_id IN (SELECT id FROM chunks WHERE file_path=?)", (fullpath,))
                c.execute("DELETE FROM chunks WHERE file_path=?", (fullpath,))
            except Exception:
                pass
            txt = self._extract_text_generic(fullpath)
            if not txt:
                return 0
            count = 0
            for ch in self._text_chunks(txt, max_chars=1200, overlap=200):
                c.execute("INSERT INTO chunks(file_path, mtime, text) VALUES(?,?,?)", (fullpath, float(mtime_ts), ch))
                cid = c.lastrowid
                v = self._hash_embedder(ch, dim=256)
                c.execute("INSERT OR REPLACE INTO embeddings(chunk_id, vec) VALUES(?,?)", (cid, self._vec_to_blob(v)))
                count += 1
            conn.commit()
            return count
        except Exception as e:
            self._append_msg(f"RAG index error en {os.path.basename(fullpath)}: {e}", "WARN")
            return 0
    # === Recuperación de contexto (RAG) ===
    def _retrieve_context(self, query: str, k: int = 6) -> str:
        # Devuelve un bloque con los k fragmentos más relevantes en toda la base RAG.
        import sqlite3, math, struct
        try:
            conn = self._rag__conn()
            c = conn.cursor()
            rows = c.execute(
                'SELECT c.id, c.text, c.file_path, e.vec '
                'FROM chunks c JOIN embeddings e ON e.chunk_id=c.id'
            ).fetchall()
            total = len(rows)
            if not rows:
                try: self._append_msg("RAG retrieve: 0 candidatos en DB.", "DEBUG")
                except Exception: pass
                return ""
            vq = self._hash_embedder(query or "", dim=256)
            def blob_to_vec(b: bytes):
                n = len(b)//4
                return list(struct.unpack('<'+'f'*n, b)) if n>0 else []
            def dot(a,b): return sum((x*y for x,y in zip(a,b)))
            def norm(a):
                s = math.sqrt(sum((x*x for x in a))) or 1.0
                return [x/s for x in a]
            vq = norm(vq)
            scored = []
            for cid, txt, fp, blob in rows:
                vv = norm(blob_to_vec(blob))
                s = dot(vq, vv)
                scored.append((s, txt, fp))
            scored.sort(reverse=True, key=lambda t: t[0])
            top = scored[:max(1,int(k))]
            try:
                self._append_msg(f"RAG retrieve: {total} candidatos, top={len(top)}.", "DEBUG")
                if top:
                    self._append_msg(f"RAG top1: {top[0][2]}", "DEBUG")
            except Exception:
                pass
            partes = []
            for i, (_, frag, fp) in enumerate(top, start=1):
                partes.append(f"[{i}] {fp}\n\"\"\"\n{frag}\n\"\"\"")
            return "\n\n".join(partes)
        except Exception as e:
            try: self._append_msg(f"RAG retrieve error: {e}", "WARN")
            except Exception: pass
            return ""

            if conn is None:
                conn = sqlite3.connect(self._db_path(), check_same_thread=False)
            self._db_rag_ensure(conn)
            c = conn.cursor()
            base = getattr(self, 'base_path', None)
            if base:
                base_str = str(base).rstrip('/\\')
                like_param = base_str + '%'
                rows = c.execute(
                    'SELECT c.id, c.text, c.file_path, e.vec FROM chunks c '
                    'JOIN embeddings e ON e.chunk_id=c.id '
                    'WHERE c.file_path LIKE ?',
                    (like_param,)
                ).fetchall()
            else:
                rows = c.execute(
                    'SELECT c.id, c.text, c.file_path, e.vec FROM chunks c '
                    'JOIN embeddings e ON e.chunk_id=c.id'
                ).fetchall()
            if not rows:
                return ""
            qv = self._hash_embedder(query or '', dim=256)
            # cosine sim
            import math as _m
            def _cos(a,b):
                num = sum(x*y for x,y in zip(a,b))
                na = _m.sqrt(sum(x*x for x in a)) or 1.0
                nb = _m.sqrt(sum(y*y for y in b)) or 1.0
                return num/(na*nb)
            scored = []
            for cid, txt, fp, blob in rows:
                v = self._blob_to_vec(blob)
                s = _cos(qv, v)
                scored.append((s, cid, txt, fp))
            scored.sort(reverse=True)
            top = scored[:max(1, int(k))]
            partes = []
            for i, (_s, _cid, txt, fp) in enumerate(top, 1):
                frag = (txt or '').strip()
                if len(frag) > 1200: frag = frag[:1200] + '…'
                partes.append(f'[{i}] {fp}\n"""\n{frag}\n"""')
            return "\n\n".join(partes)
        except Exception as e:
            self._append_msg(f'RAG retrieve error: {e}', 'WARN')
            return ""
class LLMChatDialog(tk.Toplevel):
    """Chat con modelo GGUF local usando llama-cpp-python."""
    def __init__(self, master, app):
        super().__init__(master)
        self.title("PACqui — Asistente LLM (local)")
        self.transient(master); self.resizable(True, True); self.grab_set()
        self.geometry("860x680")
        self.app = app
        self.model = None
        self.model_path = getattr(app, "llm_model_path", "") or ""
        self.stop_event = threading.Event()
        self.messages = []
        self._build_ui()
        self.protocol("WM_DELETE_WINDOW", self._on_close)
    def _build_ui(self):
        root = ttk.Frame(self, padding=10); root.pack(fill="both", expand=True)
        top = ttk.Frame(root); top.pack(fill="x")
        ttk.Label(top, text="Modelo GGUF:").pack(side="left")
        self.var_path = tk.StringVar(value=self.model_path)
        ent = ttk.Entry(top, textvariable=self.var_path); ent.pack(side="left", fill="x", expand=True, padx=6)
        ttk.Button(top, text="Elegir…", command=self._choose_model).pack(side="left")
        self.btn_cargar = ttk.Button(top, text="Cargar", command=self._load_model); self.btn_cargar.pack(side="left", padx=(6,0))
        self.lbl_status = ttk.Label(top, text="(sin cargar)"); self.lbl_status.pack(side="left", padx=(8,0))
        opts = ttk.Frame(root); opts.pack(fill="x", pady=(8,4))
        ttk.Label(opts, text="Temperatura:").pack(side="left"); self.var_temp = tk.DoubleVar(value=0.4)
        ttk.Entry(opts, width=5, textvariable=self.var_temp).pack(side="left", padx=(4,10))
        ttk.Label(opts, text="Máx. tokens:").pack(side="left"); self.var_maxtok = tk.IntVar(value=512)
        ttk.Entry(opts, width=6, textvariable=self.var_maxtok).pack(side="left", padx=(4,10))
        ttk.Label(opts, text="Contexto:").pack(side="left"); self.var_ctx = tk.IntVar(value=8192)
        ttk.Entry(opts, width=7, textvariable=self.var_ctx).pack(side="left", padx=(4,10))
        sysf = ttk.LabelFrame(root, text="System"); sysf.pack(fill="x")
        self.txt_sys = ScrolledText(sysf, height=3, wrap="word"); self.txt_sys.insert("1.0", "Soy PACqui, tu asistente para las dudas que tengas acerca de la PAC \n."); self.txt_sys.pack(fill="x")
        chatf = ttk.LabelFrame(root, text="Conversación"); chatf.pack(fill="both", expand=True, pady=(8,0))
        self.txt_chat = ScrolledText(chatf, wrap="word"); self.txt_chat.pack(fill="both", expand=True); self.txt_chat.configure(state="disabled")
        bot = ttk.Frame(root); bot.pack(fill="x", pady=(8,0))
        self.var_user = tk.StringVar()
        ent_u = ttk.Entry(bot, textvariable=self.var_user); ent_u.pack(side="left", fill="x", expand=True); ent_u.bind("<Return>", lambda e: self._send())
        ttk.Button(bot, text="Enviar", command=self._send).pack(side="left", padx=6)
        self.btn_stop = ttk.Button(bot, text="Detener", command=self._stop, state="disabled"); self.btn_stop.pack(side="left")
        if self.model_path and Path(self.model_path).exists(): self.after(250, self._load_model)
    def _choose_model(self):
        path = filedialog.askopenfilename(title="Selecciona el archivo .gguf", filetypes=[("GGUF","*.gguf"),("Todos","*.*")])
        if path: self.var_path.set(path)
    def _append_chat(self, who, text):
        self.txt_chat.configure(state="normal")
        self.txt_chat.insert("end", f"{who}: ", ("who",)); self.txt_chat.insert("end", (text or "").strip()+"\n")
        self.txt_chat.tag_configure("who", font=("Segoe UI", 9, "bold")); self.txt_chat.see("end"); self.txt_chat.configure(state="disabled")
    def _set_status(self, s):
        try: self.lbl_status.configure(text=s); self.update_idletasks()
        except Exception: pass
    def _load_model(self):
        path = self.var_path.get().strip()
        if not path or not Path(path).exists():
            messagebox.showwarning(APP_NAME, "Selecciona primero un archivo .gguf válido."); return
        self.model_path = path; self._set_status("Cargando…"); self.btn_cargar.configure(state="disabled")
        threading.Thread(target=self._worker_load_model, daemon=True).start()
    def _worker_load_model(self):
        try:
            try:
                from llama_cpp import Llama
            except Exception:
                self.after(0, lambda: messagebox.showerror(APP_NAME, "No está instalado 'llama-cpp-python'.\\n\\nInstala:\\nCPU: pip install llama-cpp-python --extra-index-url https://abetlen.github.io/llama-cpp-python/whl/cpu\\nCUDA: pip install llama-cpp-python --extra-index-url https://abetlen.github.io/llama-cpp-python/whl/cu124"))
                self.after(0, lambda: self._set_status("(falta instalar)")); self.after(0, lambda: self.btn_cargar.configure(state="normal")); return
            ctx = int(self.var_ctx.get() or 4096)
            self.model = Llama(model_path=self.model_path, n_ctx=ctx, n_gpu_layers=-1, chat_format="mistral-instruct")
            self.app.llm_model_path = self.model_path
            try: self.app._save_config()
            except Exception: pass
            self.after(0, lambda: self._set_status("Modelo cargado ✓"))
        except Exception as e:
            self.after(0, lambda: self._set_status("(error)"))
            self.after(0, lambda: messagebox.showerror(APP_NAME, f"Error cargando el modelo:\\n{e}"))
        finally:
            self.after(0, lambda: self.btn_cargar.configure(state="normal"))
    def _stop(self): self.stop_event.set()
    
    def _send(self):
        user = (self.var_user.get() or "").strip()
        if not user: return
        self.var_user.set(""); self._append_chat("Tú", user)
        if self.model is None:
            self._append_chat("PACqui", "Primero carga el modelo (botón Cargar)."); return
        # Recuperar contexto acotado a la carpeta base activa
        try:
            ctx = getattr(self.app, '_retrieve_context', lambda *_a, **_k: '')(user, k=6)
        except Exception as e:
            self._append_chat("PACqui", f"Error recuperando contexto: {e}"); return
        if not ctx:
            self._append_chat("PACqui", "No está en la base (vuelve a escanear la carpeta base o afina la búsqueda)."); return
        # Prompt endurecido
        sys_prompt = "Eres PACqui, asistente de documentación para la PAC. Responde SOLO usando el CONTEXTO proporcionado. Si la información no aparece, responde exactamente: 'No está en la base'. Cita los fragmentos como [n] si procede."
        self.messages = [{"role":"system","content":sys_prompt},
                         {"role":"user","content": f"### CONTEXTO RELEVANTE (solo documentos de la carpeta base actual):\n{ctx}\n\n### PREGUNTA:\n{user}"}]
        self.stop_event.clear(); self.btn_stop.configure(state="normal")
        threading.Thread(target=self._worker_chat_stream, daemon=True).start()

    def _worker_chat_stream(self):
        try:
            temp = float(self.var_temp.get() or 0.3); max_t = int(self.var_maxtok.get() or 512); out=[]
            for chunk in self.model.create_chat_completion(messages=self.messages, temperature=temp, max_tokens=max_t, stream=True):
                if self.stop_event.is_set(): break
                try: delta = chunk["choices"][0]["delta"].get("content","")
                except Exception: delta = ""
                if delta: out.append(delta); self.after(0, lambda t=delta: self._append_stream_text(t))
            final = "".join(out).strip()
            if final: self.messages.append({"role":"assistant","content":final}); self.after(0, lambda: self._append_stream_text("\n", end_turn=True))
        except Exception as e:
            self.after(0, lambda: messagebox.showerror(APP_NAME, f"Error en inferencia:\\n{e}"))
        finally:
            self.after(0, lambda: self.btn_stop.configure(state="normal"))
    def _append_stream_text(self, txt, end_turn=False):
        self.txt_chat.configure(state="normal")
        if not hasattr(self, "_in_stream"): self._in_stream = False
        if not self._in_stream: self.txt_chat.insert("end", "PACqui:\n", ("who",)); self._in_stream = True
        self.txt_chat.insert("end", txt)
        if end_turn: self.txt_chat.insert("end", "\n"); self._in_stream = False
        self.txt_chat.tag_configure("who", font=("Segoe UI", 9, "bold")); self.txt_chat.see("end"); self.txt_chat.configure(state="disabled")
    def _on_close(self):
        self.stop_event.set(); self.destroy()


def main():
    root = tk.Tk()
    root.title(f"{APP_NAME} – v{APP_VERSION}")
    root.geometry("1400x820")
    try:
        root.call("source", "azure.tcl")
        ttk.Style().theme_use("azure")
    except Exception:
        pass

    app = OrganizadorFrame(root)
    app.pack(fill="both", expand=True)
    app._append_msg("Selecciona la carpeta base y escanea. Usa 'Exportar ▾' para Excel (visibles o todo).", "INFO")
    root.bind('<F1>', lambda e: app.cmd_ayuda())
    root.mainloop()


if __name__ == "__main__":
    main()